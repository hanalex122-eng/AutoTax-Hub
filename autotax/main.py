import asyncio
import logging
import os
from datetime import datetime
from typing import Optional, List
from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Query, Body, Request
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, HTMLResponse
from pydantic import BaseModel
import io

from autotax.ocr import extract_text, extract_text_and_qr
from autotax.parser import parse_invoice
from autotax.db import init_db, save_invoice, SessionLocal
from autotax.models import Invoice, User, CashEntry, UserCompany
from autotax.auth import hash_password, verify_password, create_token, create_access_token, create_refresh_token, decode_token, get_current_user

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("autotax")

limiter = Limiter(key_func=get_remote_address)

app = FastAPI(
    title="AutoTax-HUB",
    version="5.5.0",
)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

_allowed_origins = os.getenv(
    "ALLOWED_ORIGINS",
    "https://web-production-489ac.up.railway.app,https://app.autotaxhub.de,http://localhost:3000,http://localhost:5173"
).split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in _allowed_origins if o.strip()],
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
    max_age=3600,
)


@app.middleware("http")
async def security_headers(request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(self), microphone=()"
    response.headers["X-Data-Retention"] = "none"
    return response


def ok_list(items, total):
    return {"success": True, "items": items, "total": total}


def err(status: int, msg: str):
    raise HTTPException(status_code=status, detail={"success": False, "error": msg})


def safe_str(val, default=""):
    return val if val is not None else default


def safe_float(val, default=0.0):
    return val if val is not None else default


def safe_vat_rate(val):
    return val if val else "0%"


def safe_vendor(val):
    return val if val else "Unbekannt"


def safe_category(val):
    return val if val else "other"


def safe_invoice_type(val):
    return val if val in ("income", "expense") else "expense"


def safe_date_str(val):
    if not val:
        return ""
    return val


def parse_vat_rate_float(vat_rate_str):
    try:
        return float((vat_rate_str or "0").replace("%", ""))
    except (ValueError, TypeError):
        return 0.0


def calc_vat(gross, vat_rate_str):
    if not gross:
        return 0.0
    rate = parse_vat_rate_float(vat_rate_str)
    if rate <= 0:
        return 0.0
    return round(gross * rate / (100 + rate), 2)


def _fuzzy_match(a: str, b: str, threshold: float = 0.75) -> bool:
    if not a or not b:
        return False
    a, b = a.lower().replace(" ", ""), b.lower().replace(" ", "")
    if a == b or a in b or b in a:
        return True
    common = sum(1 for c in a if c in b)
    return common / max(len(a), len(b)) >= threshold


def parse_date_str_to_datetime(date_str):
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str)
    except ValueError:
        pass
    try:
        return datetime.strptime(date_str, "%d.%m.%Y")
    except ValueError:
        pass
    return None


def auto_create_cash_entry(invoice_id: int, user_id: int, data: dict):
    """Create a CashEntry automatically when an invoice is uploaded."""
    db = SessionLocal()
    try:
        # Skip if already synced
        existing = db.query(CashEntry).filter(CashEntry.invoice_id == invoice_id, CashEntry.user_id == user_id).first()
        if existing:
            return
        # Parse date safely
        date_val = None
        date_str = data.get("date") or ""
        if date_str:
            date_val = parse_date_str_to_datetime(date_str)
        if not date_val:
            date_val = datetime.now()
        entry = CashEntry(
            user_id=user_id,
            description=f"Rechnung: {data.get('vendor') or 'Unbekannt'}",
            vendor=data.get("vendor") or "Unbekannt",
            gross_amount=float(data.get("total_amount") or 0),
            vat_amount=float(data.get("vat_amount") or 0),
            vat_rate=data.get("vat_rate") or "0%",
            entry_type="expense",
            category=data.get("category") or "other",
            payment_method=data.get("payment_method") or "",
            reference=data.get("invoice_number") or f"INV-{invoice_id}",
            notes=f"Auto-sync from invoice #{invoice_id}",
            is_reconciled=False,
            invoice_id=invoice_id,
            date=date_val,
        )
        db.add(entry)
        db.commit()
        logger.info("Auto-synced invoice %s to cash_entries", invoice_id)
    except Exception:
        db.rollback()
        logger.exception("Auto cash entry creation failed for invoice %s", invoice_id)
    finally:
        db.close()


# --- ADDED START: quick entity extractors for invoice_to_dict ---
import re as _re_global

def _extract_first_iban(text):
    m = _re_global.search(r"\b([A-Z]{2}\s?\d{2}\s?(?:\d{4}\s?){2,7}\d{1,4})\b", text.upper())
    return m.group(1).replace(" ", "") if m else ""

def _extract_first_phone(text):
    m = _re_global.search(r"(?:tel\.?|fon|phone|fax)\s*:?\s*([\d\s/\-+]{6,20})", text, _re_global.IGNORECASE)
    return m.group(1).strip() if m else ""

def _extract_first_address(text):
    m = _re_global.search(r"(\d{4,5}\s+[A-ZÄÖÜ][a-zäöüß]{2,}(?:\s+[A-ZÄÖÜ][a-zäöüß]{2,})?)", text)
    return m.group(1).strip() if m else ""
# --- ADDED END ---

def invoice_to_dict(i):
    return {
        "id": i.id,
        "vendor": safe_vendor(i.vendor),
        "invoice_number": safe_str(i.invoice_number),
        "invoice_type": safe_invoice_type(i.invoice_type),
        "total_amount": safe_float(i.total_amount),
        "vat_amount": safe_float(i.vat_amount),
        "vat_rate": safe_vat_rate(i.vat_rate),
        "date": safe_date_str(i.date),
        "payment_method": safe_str(i.payment_method),
        "category": safe_category(i.category),
        "processed": i.processed or False,
        "created_at": i.created_at.strftime("%Y-%m-%dT%H:%M:%S") if i.created_at else "",
        "ocr_snippet": (i.raw_text or "")[:200],
        "konto": _DATEV_KONTO_MAP.get(safe_category(i.category), "6800") if safe_invoice_type(i.invoice_type) == "expense" else _DATEV_KONTO_MAP_INCOME.get(safe_category(i.category), "8400"),
        "has_original": bool(i.file_data and len(i.file_data) > 0) if i.file_data else False,
        # --- ADDED: vendor details from raw_text ---
        "vendor_iban": _extract_first_iban(i.raw_text or ""),
        "vendor_phone": _extract_first_phone(i.raw_text or ""),
        "vendor_address": _extract_first_address(i.raw_text or ""),
    }


def cash_entry_to_dict(e):
    return {
        "id": e.id,
        "description": safe_str(e.description),
        "vendor": safe_vendor(e.vendor),
        "gross_amount": safe_float(e.gross_amount),
        "vat_amount": safe_float(e.vat_amount),
        "vat_rate": safe_vat_rate(e.vat_rate),
        "entry_type": safe_invoice_type(e.entry_type),
        "category": safe_category(e.category),
        "payment_method": safe_str(e.payment_method),
        "reference": safe_str(e.reference),
        "notes": safe_str(e.notes),
        "is_reconciled": e.is_reconciled or False,
        "invoice_id": e.invoice_id,
        "date": e.date.strftime("%Y-%m-%d") if e.date else "",
        "created_at": e.created_at.strftime("%Y-%m-%dT%H:%M:%S") if e.created_at else "",
    }


@app.on_event("startup")
def startup():
    init_db()


@app.get("/health")
def health():
    ocr_key = os.getenv("OCR_API_KEY", "")
    return {"status": "ok", "version": "5.5.5", "ocr_configured": bool(ocr_key), "ocr_key_len": len(ocr_key)}


@app.get("/manifest.json")
def pwa_manifest():
    return {
        "name": "AutoTax-HUB",
        "short_name": "AutoTax",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#050a12",
        "theme_color": "#10b981",
        "description": "Automatische Rechnungserkennung & Buchhaltung",
        "icons": [
            {"src": "data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><rect fill='%23050a12' width='100' height='100' rx='20'/><text x='50' y='65' font-size='50' text-anchor='middle' fill='%2310b981' font-family='sans-serif' font-weight='bold'>AT</text></svg>", "sizes": "192x192", "type": "image/svg+xml"},
        ],
    }


from fastapi.responses import Response as RawResponse


@app.get("/sw.js")
def service_worker():
    return RawResponse(content="self.addEventListener('fetch',e=>{});", media_type="application/javascript")


@app.get("/invoices/{invoice_id}/pdf")
def generate_invoice_pdf(invoice_id: int, user: dict = Depends(get_current_user)):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as pdf_canvas
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor
    except ImportError:
        raise HTTPException(status_code=501, detail="PDF-Generierung nicht verfügbar (reportlab fehlt)")

    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            raise HTTPException(status_code=404, detail="Rechnung nicht gefunden")

        companies = db.query(UserCompany).filter(UserCompany.user_id == user["sub"]).all()
        company_name = companies[0].company_name if companies else "Meine Firma"
        u = db.query(User).filter(User.id == user["sub"]).first()

        buf = io.BytesIO()
        c = pdf_canvas.Canvas(buf, pagesize=A4)
        w, h = A4

        c.setFillColor(HexColor("#1a2d4a"))
        c.setFont("Helvetica-Bold", 22)
        c.drawString(2*cm, h-2.5*cm, company_name)
        c.setFillColor(HexColor("#00e5a0"))
        c.setFont("Helvetica", 9)
        c.drawString(2*cm, h-3*cm, f"E-Mail: {u.email if u else ''}")

        c.setFillColor(HexColor("#1a2d4a"))
        c.setFont("Helvetica-Bold", 16)
        typ = "RECHNUNG" if inv.invoice_type == "income" else "BELEG"
        c.drawString(2*cm, h-4.5*cm, typ)
        c.setFont("Helvetica", 11)
        c.drawString(12*cm, h-4.5*cm, f"Nr: {inv.invoice_number or f'RE-{inv.id}'}")
        c.drawString(12*cm, h-5.1*cm, f"Datum: {inv.date or 'k.A.'}")

        c.setFont("Helvetica-Bold", 11)
        c.drawString(2*cm, h-6*cm, "An:" if inv.invoice_type == "income" else "Von:")
        c.setFont("Helvetica", 11)
        c.drawString(2*cm, h-6.6*cm, inv.vendor or "Unbekannt")

        y = h - 8.5*cm
        c.setFillColor(HexColor("#1a2d4a"))
        c.rect(2*cm, y, 17*cm, 0.8*cm, fill=1)
        c.setFillColor(HexColor("#ffffff"))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(2.2*cm, y+0.25*cm, "Beschreibung")
        c.drawString(10*cm, y+0.25*cm, "Kategorie")
        c.drawString(13*cm, y+0.25*cm, "MwSt")
        c.drawString(16*cm, y+0.25*cm, "Betrag")

        y -= 0.8*cm
        c.setFillColor(HexColor("#1a2d4a"))
        c.setFont("Helvetica", 10)
        c.drawString(2.2*cm, y+0.25*cm, inv.vendor or "Position 1")
        c.drawString(10*cm, y+0.25*cm, inv.category or "other")
        c.drawString(13*cm, y+0.25*cm, f"{inv.vat_rate or '19%'}")
        c.drawString(16*cm, y+0.25*cm, f"EUR {inv.total_amount or 0:.2f}")
        c.line(2*cm, y, 19*cm, y)

        y -= 1.5*cm
        netto = (inv.total_amount or 0) - (inv.vat_amount or 0)
        c.setFont("Helvetica", 10)
        c.drawRightString(19*cm, y, f"Netto: EUR {netto:.2f}")
        y -= 0.5*cm
        c.drawRightString(19*cm, y, f"MwSt ({inv.vat_rate or '19%'}): EUR {inv.vat_amount or 0:.2f}")
        y -= 0.6*cm
        c.setFont("Helvetica-Bold", 12)
        c.drawRightString(19*cm, y, f"Gesamtbetrag: EUR {inv.total_amount or 0:.2f}")

        if hasattr(u, 'is_kleinunternehmer') and getattr(u, 'is_kleinunternehmer', False):
            y -= 1.5*cm
            c.setFont("Helvetica-Oblique", 8)
            c.setFillColor(HexColor("#7a8ba8"))
            c.drawString(2*cm, y, "Gemäß §19 UStG wird keine Umsatzsteuer berechnet.")

        c.setFillColor(HexColor("#7a8ba8"))
        c.setFont("Helvetica", 7)
        c.drawString(2*cm, 1.5*cm, f"Erstellt mit AutoTax-HUB | {company_name} | {u.email if u else ''}")
        c.drawString(2*cm, 1*cm, "Automatisch erstellt. Alle Angaben ohne Gewähr. Keine Steuerberatung.")

        c.save()
        buf.seek(0)

        filename = f"{typ}_{inv.invoice_number or inv.id}.pdf"
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})
    finally:
        db.close()


@app.get("/", response_class=HTMLResponse)
async def serve_frontend():
    index_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "index.html")
    with open(index_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read(), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


@app.get("/app", response_class=HTMLResponse)
async def serve_frontend_app():
    index_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "index.html")
    with open(index_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read(), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


# --- ADDED START: Landing page ---
@app.get("/landing", response_class=HTMLResponse)
async def serve_landing_page():
    lp = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "landing-new.html")
    with open(lp, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read(), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})
# --- ADDED END ---

# --- ADDED START: Split-view editor page ---
@app.get("/editor", response_class=HTMLResponse)
async def serve_editor_page():
    ep = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "editor.html")
    with open(ep, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read(), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})
# --- ADDED END ---

# --- ADDED START: Beleg entry page ---
@app.get("/beleg", response_class=HTMLResponse)
async def serve_beleg_page():
    beleg_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "beleg.html")
    with open(beleg_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read(), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})
# --- ADDED END ---


@app.get("/agb", response_class=HTMLResponse)
def agb_page():
    return HTMLResponse(content="""<!DOCTYPE html><html lang="de"><head><meta charset="UTF-8"><title>AGB — AutoTax-HUB</title>
<style>body{font-family:'DM Sans',sans-serif;max-width:800px;margin:40px auto;padding:20px;background:#050a12;color:#e8edf5;line-height:1.8}
h1{color:#10b981;font-size:28px}h2{color:#00a8cc;margin-top:30px;font-size:18px}strong{color:#f59e0b}
a{color:#10b981}p{margin:12px 0}</style></head><body>
<h1>Allgemeine Geschäftsbedingungen</h1>
<p><em>AutoTax-HUB — Stand: März 2026</em></p>
<h2>§1 Geltungsbereich</h2>
<p>Diese AGB gelten für die Nutzung von AutoTax-HUB, einem Software-Werkzeug zur automatischen Erfassung und Verwaltung von Belegen und Kassenbüchern.</p>
<h2>§2 Leistungsbeschreibung</h2>
<p>AutoTax-HUB bietet: OCR-Erkennung von Belegen, automatische Kategorisierung, Kassenbuchführung, Export in verschiedene Formate (CSV, DATEV, Excel, JSON), sowie eine EÜR-Übersicht.</p>
<h2>§3 Keine Steuerberatung</h2>
<p>AutoTax-HUB ist ein Software-Werkzeug. Der Dienst stellt <strong>KEINE Steuerberatung</strong> dar. Die automatische Erkennung kann Fehler enthalten. Der Nutzer ist <strong>allein verantwortlich</strong> für die Überprüfung aller Daten und die Richtigkeit seiner Buchhaltung. Für die Steuererklärung wird dringend empfohlen, einen Steuerberater zu konsultieren.</p>
<h2>§4 Haftungsbeschränkung</h2>
<p>Der Anbieter haftet <strong>nicht</strong> für: falsch erkannte Beträge, fehlerhafte MwSt-Berechnungen, steuerliche Nachteile, verlorene Daten oder sonstige Schäden, die aus der Nutzung des Dienstes entstehen. Die Haftung ist auf den vom Nutzer gezahlten Betrag der letzten 12 Monate beschränkt.</p>
<h2>§5 Prüfungspflicht</h2>
<p>Der Nutzer ist <strong>verpflichtet</strong>, alle automatisch erkannten Daten (Beträge, Lieferanten, Kategorien, MwSt-Sätze, Datumsangaben) vor Verwendung zu überprüfen und ggf. zu korrigieren.</p>
<h2>§6 Datenschutz</h2>
<p>Daten werden DSGVO-konform in der EU gespeichert. Originalbilder werden nach der OCR-Verarbeitung nicht dauerhaft auf dem Server gespeichert. Personenbezogene Daten werden nicht an Dritte weitergegeben. Details siehe <a href="/datenschutz">Datenschutzerklärung</a>.</p>
<h2>§7 Beta-Phase</h2>
<p>AutoTax-HUB befindet sich in der <strong>Beta-Phase</strong>. Funktionen können sich ändern. Es wird keine Garantie für ununterbrochene Verfügbarkeit gegeben.</p>
<h2>§8 Kündigung</h2>
<p>Das Nutzerkonto kann jederzeit gelöscht werden. Alle Daten werden dabei unwiderruflich entfernt.</p>
<p style="margin-top:40px;color:#64748b;font-size:13px">© 2026 AutoTax-HUB — Alle Rechte vorbehalten.</p>
</body></html>""")


@app.post("/admin/reset-password")
def admin_reset_password(body: dict = Body(...)):
    email = body.get("email")
    new_password = body.get("new_password")
    if not email or not new_password:
        err(400, "email and new_password required")
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == email).first()
        if not user:
            err(404, "User not found")
        user.hashed_password = hash_password(new_password)
        db.commit()
        return {"success": True, "message": f"Password reset for {email}"}
    finally:
        db.close()


@app.post("/admin/reparse")
def admin_reparse(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        invoices = db.query(Invoice).all()
        count = 0
        for inv in invoices:
            if not inv.raw_text:
                continue
            parsed = parse_invoice(inv.raw_text)
            inv.total_amount = parsed["total_amount"]
            inv.vat_amount = parsed["vat_amount"]
            inv.vat_rate = parsed["vat_rate"]
            inv.vendor = parsed["vendor"]
            inv.category = parsed["category"]
            inv.date = parsed["date"]
            count += 1
        db.commit()
        return {"status": "done", "count": count}
    except Exception:
        db.rollback()
        logger.exception("Reparse failed")
        err(500, "Reparse failed")
    finally:
        db.close()


# DATEV Konto mapping (Ausgaben)
_DATEV_KONTO_MAP = {
    "food": "6800", "groceries": "6800", "restaurant": "6640",
    "fuel": "6670", "transport": "6673",
    "office": "6815", "software": "6815", "subscription": "6815",
    "telecom": "6805", "shipping": "6810",
    "electronics": "6800", "shopping": "6800",
    "insurance": "6400", "health": "6800", "medical": "6800",
    "home": "6800", "clothing": "6800",
    "other": "6800",
}
# DATEV Konto mapping (Einnahmen)
_DATEV_KONTO_MAP_INCOME = {
    "other": "8400", "food": "8400", "electronics": "8400",
    "software": "8400", "shopping": "8400",
}

ALLOWED_TYPES = {"application/pdf", "image/jpeg", "image/jpg", "image/png", "image/tiff", "image/webp", "image/heic", "image/heif", "application/zip", "application/x-zip-compressed"}
MAX_FILE_SIZE = 10 * 1024 * 1024

# Magic bytes for file type validation (prevents fake content_type)
_MAGIC_BYTES = {
    b"\xff\xd8\xff": "image/jpeg",
    b"\x89PNG": "image/png",
    b"%PDF": "application/pdf",
    b"II\x2a\x00": "image/tiff",  # little-endian TIFF
    b"MM\x00\x2a": "image/tiff",  # big-endian TIFF
    b"RIFF": "image/webp",        # WebP starts with RIFF
    b"PK\x03\x04": "application/zip",  # ZIP archive
}


def _validate_file_magic(content: bytes, claimed_type: str) -> bool:
    """Check if file content matches claimed MIME type via magic bytes."""
    if not content or len(content) < 4:
        return False
    # HEIC/HEIF have complex headers — trust content_type for those
    if "heic" in claimed_type or "heif" in claimed_type:
        return True
    for magic, mime in _MAGIC_BYTES.items():
        if content[:len(magic)] == magic:
            return True
    return False


# ============================================================
# AUTH
# ============================================================

class AuthRequest(BaseModel):
    email: str
    password: str


class RegisterRequest(BaseModel):
    email: str
    password: str
    full_name: Optional[str] = None
    company_name: Optional[str] = None


@app.post("/auth/register")
@limiter.limit("3/minute")
def register(request: Request, body: RegisterRequest):
    if len(body.password) < 8:
        err(400, "Password must be at least 8 characters")
    if not any(c.isupper() for c in body.password):
        err(400, "Password must contain at least 1 uppercase letter")
    if not any(c.isdigit() for c in body.password):
        err(400, "Password must contain at least 1 digit")
    db = SessionLocal()
    try:
        if db.query(User).filter(User.email == body.email).first():
            err(400, "Email already registered")
        try:
            user = User(email=body.email, hashed_password=hash_password(body.password), full_name=body.full_name, plan="early")
        except Exception:
            user = User(email=body.email, hashed_password=hash_password(body.password), full_name=body.full_name)
        db.add(user)
        db.commit()
        db.refresh(user)
        # Auto-create company (optional — don't fail registration if this fails)
        comp_name = ""
        try:
            comp_name = (body.company_name or "").strip()
            if not comp_name:
                comp_name = (body.full_name or "").strip()
            if not comp_name:
                comp_name = body.email.split("@")[0].strip()
            if comp_name:
                company = UserCompany(user_id=user.id, company_name=comp_name)
                db.add(company)
                db.commit()
        except Exception:
            logger.warning("Could not create company for %s — table may not exist yet", body.email)
        logger.info("User registered: %s (company: %s)", body.email, comp_name)
        token = create_token(user.id, user.email)
        return {"success": True, "token": token, "email": user.email}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Registration error")
        err(500, "Registration failed")
    finally:
        db.close()


@app.post("/auth/login")
@limiter.limit("5/minute")
def login(request: Request, body: AuthRequest):
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == body.email).first()
        if not user or not verify_password(body.password, user.hashed_password):
            logger.warning("Failed login: %s", body.email)
            err(401, "Invalid email or password")
        logger.info("User logged in: %s", body.email)
        token = create_access_token(user.id, user.email)
        refresh = create_refresh_token(user.id, user.email)
        return {"success": True, "token": token, "refresh_token": refresh, "email": user.email}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Login error")
        err(500, "Login failed")
    finally:
        db.close()


@app.post("/auth/refresh")
def refresh_token_endpoint(body: dict = Body(...)):
    refresh = body.get("refresh_token", "")
    if not refresh:
        err(400, "refresh_token required")
    try:
        data = decode_token(refresh, expected_type="refresh")
    except HTTPException:
        err(401, "Invalid or expired refresh token")
    new_access = create_access_token(data["sub"], data["email"])
    new_refresh = create_refresh_token(data["sub"], data["email"])
    return {"success": True, "token": new_access, "refresh_token": new_refresh}


class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str


@app.post("/auth/change-password")
def change_password(body: ChangePasswordRequest, user: dict = Depends(get_current_user)):
    if len(body.new_password) < 8:
        err(400, "Neues Passwort muss mindestens 8 Zeichen haben")
    if not any(c.isupper() for c in body.new_password):
        err(400, "Neues Passwort muss mindestens 1 Großbuchstaben enthalten")
    if not any(c.isdigit() for c in body.new_password):
        err(400, "Neues Passwort muss mindestens 1 Zahl enthalten")
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user["sub"]).first()
        if not u or not verify_password(body.old_password, u.hashed_password):
            err(401, "Altes Passwort ist falsch")
        u.hashed_password = hash_password(body.new_password)
        db.commit()
        return {"success": True, "message": "Passwort erfolgreich geändert"}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Password change failed")
        err(500, "Passwort-Änderung fehlgeschlagen")
    finally:
        db.close()


@app.post("/auth/reset-password")
@limiter.limit("3/minute")
def reset_password(request: Request, body: dict = Body(...)):
    """Send password reset — for now just verify email exists and return token."""
    email = body.get("email", "").strip().lower()
    if not email:
        err(400, "E-Mail erforderlich")
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.email == email).first()
        if not u:
            # Don't reveal if email exists
            return {"success": True, "message": "Falls ein Konto existiert, wurde ein Reset-Link gesendet."}
        # Generate reset token (valid 1 hour)
        reset_token = create_access_token(u.id, u.email)
        logger.info("Password reset requested for %s", email)
        # TODO: Send email with reset link. For now return token directly.
        return {"success": True, "message": "Reset-Token generiert. Bitte kontaktiere den Admin.", "reset_token": reset_token}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Password reset failed")
        err(500, "Reset fehlgeschlagen")
    finally:
        db.close()


# ============================================================
# INVOICES: UPLOAD
# ============================================================


@app.post("/invoices/upload-erechnung")
async def upload_erechnung(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Import XRechnung / ZUGFeRD / Factur-X XML e-invoice."""
    import xml.etree.ElementTree as ET
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        err(400, "Datei zu groß (max 5MB)")
    text = content.decode("utf-8", errors="ignore")

    # XXE protection: strip DOCTYPE declarations before parsing
    import re as _re
    text_safe = _re.sub(r'<!DOCTYPE[^>]*>', '', text, flags=_re.IGNORECASE | _re.DOTALL)
    text_safe = _re.sub(r'<!ENTITY[^>]*>', '', text_safe, flags=_re.IGNORECASE | _re.DOTALL)

    root = None
    try:
        root = ET.fromstring(text_safe)
    except ET.ParseError:
        pass

    if root is None:
        err(400, "Keine gültige XML/E-Rechnung")

    # Parse with namespace-agnostic approach
    def _find(el, tags):
        for tag in tags:
            for child in el.iter():
                if tag.lower() in child.tag.lower():
                    if child.text and child.text.strip():
                        return child.text.strip()
        return ""

    vendor = _find(root, ["PartyName", "Name", "SellerTradeParty"])
    invoice_number = _find(root, ["InvoiceNumber", "DocumentNumber"]) or _find(root, ["ID"])
    date_str = _find(root, ["IssueDate", "DateTimeString", "InvoiceDate"])
    total_str = _find(root, ["PayableAmount", "TaxInclusiveAmount", "GrandTotalAmount", "DuePayableAmount"])
    tax_str = _find(root, ["TaxAmount", "TaxTotalAmount"])
    vat_rate_str = _find(root, ["Percent", "RateApplicablePercent", "CategoryCode"])

    total = 0.0
    try:
        total = float(total_str.replace(",", "."))
    except (ValueError, AttributeError):
        pass

    tax = 0.0
    try:
        tax = float(tax_str.replace(",", "."))
    except (ValueError, AttributeError):
        pass

    vat_rate = "19%"
    try:
        r = float(vat_rate_str.replace(",", "."))
        if 0 < r <= 30:
            vat_rate = f"{r}%"
    except (ValueError, AttributeError):
        pass

    if not vendor:
        vendor = "E-Rechnung"

    # Use existing category detection
    try:
        from autotax.parser import detect_category
        category = detect_category(vendor, text)
    except Exception:
        category = "other"

    # Save invoice
    db = SessionLocal()
    try:
        inv = Invoice(
            user_id=user["sub"],
            filename=file.filename or "e-rechnung.xml",
            vendor=vendor,
            total_amount=total,
            vat_amount=tax if tax > 0 else round(total * 19 / 119, 2),
            vat_rate=vat_rate,
            date=date_str,
            raw_text=text[:2000],
            invoice_type="expense",
            invoice_number=invoice_number,
            payment_method="",
            category=category,
            processed=True,
            file_data=content,
            file_content_type=file.content_type or "application/xml",
        )
        db.add(inv)
        db.commit()
        db.refresh(inv)
        auto_create_cash_entry(inv.id, user["sub"], {
            "vendor": vendor, "total_amount": total,
            "vat_amount": tax, "vat_rate": vat_rate,
            "date": date_str, "category": category,
        })
        konto = _DATEV_KONTO_MAP.get(category, "6800")
        return {
            "success": True,
            "id": inv.id,
            "vendor": vendor,
            "total_amount": total,
            "vat_amount": tax,
            "vat_rate": vat_rate,
            "date": date_str,
            "invoice_number": invoice_number,
            "category": category,
            "konto": konto,
            "message": "E-Rechnung erfolgreich importiert — Automatisch kategorisiert ✔",
        }
    except Exception:
        db.rollback()
        logger.exception("E-Rechnung import failed")
        err(500, "E-Rechnung Import fehlgeschlagen")
    finally:
        db.close()


@app.post("/invoices/create-rechnung")
def create_rechnung(body: dict = Body(...), user: dict = Depends(get_current_user)):
    """Create a manual outgoing invoice (Einnahme)."""
    db = SessionLocal()
    try:
        betrag = float(body.get("betrag", 0))
        mwst_satz = body.get("mwst_satz", "19%")
        rate = float(mwst_satz.replace("%", "").replace(",", ".").strip() or "19")
        mwst_betrag = float(body.get("mwst_betrag", 0)) or round(betrag * rate / (100 + rate), 2)
        inv = Invoice(
            user_id=user["sub"], filename="rechnung-erstellt",
            vendor=body.get("kunde", ""), total_amount=betrag,
            vat_amount=mwst_betrag, vat_rate=mwst_satz,
            date=body.get("datum", ""), raw_text="Manuell erstellte Rechnung",
            invoice_type="income", invoice_number=body.get("rechnung_nr", ""),
            payment_method=body.get("zahlungsart", ""),
            category=body.get("kategorie", "service"), processed=True,
        )
        db.add(inv)
        db.commit()
        db.refresh(inv)
        auto_create_cash_entry(inv.id, user["sub"], {
            "vendor": body.get("kunde", ""), "total_amount": betrag,
            "vat_amount": mwst_betrag, "vat_rate": mwst_satz,
            "date": body.get("datum", ""), "category": "service",
            "invoice_type": "income",
        })
        return {"success": True, "id": inv.id, "invoice_number": inv.invoice_number}
    except HTTPException:
        raise
    except Exception:
        db.rollback()
        logger.exception("Create Rechnung failed")
        err(500, "Rechnung erstellen fehlgeschlagen")
    finally:
        db.close()


@app.post("/account/kleinunternehmer")
def toggle_kleinunternehmer(body: dict = Body(...), user: dict = Depends(get_current_user)):
    """Toggle Kleinunternehmerregelung §19 UStG."""
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user["sub"]).first()
        if not u:
            err(404, "User not found")
        val = body.get("enabled", False)
        # Store in full_name field as prefix (backward compatible — no schema change needed)
        if val and not (u.full_name or "").startswith("[KU]"):
            u.full_name = f"[KU] {u.full_name or ''}"
        elif not val and (u.full_name or "").startswith("[KU]"):
            u.full_name = (u.full_name or "").replace("[KU] ", "").strip()
        db.commit()
        return {"success": True, "kleinunternehmer": val}
    finally:
        db.close()


@app.get("/account/kleinunternehmer")
def get_kleinunternehmer(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user["sub"]).first()
        is_ku = (u.full_name or "").startswith("[KU]") if u else False
        return {"kleinunternehmer": is_ku}
    finally:
        db.close()


@app.post("/invoices/create")
def create_invoice_manual(body: dict = Body(...), user: dict = Depends(get_current_user)):
    """Create invoice from JSON (for cross-page transfer). Skips duplicates."""
    db = SessionLocal()
    try:
        # Duplicate check
        dup = db.query(Invoice).filter(
            Invoice.user_id == user["sub"],
            Invoice.vendor == (body.get("vendor") or "Manual"),
            Invoice.total_amount == float(body.get("total_amount") or 0),
        ).first()
        if dup:
            return {"success": True, "id": dup.id, "message": "already exists"}
        inv = Invoice(
            user_id=user["sub"],
            filename=None,
            vendor=body.get("vendor") or "Manual",
            total_amount=float(body.get("total_amount") or 0),
            vat_amount=float(body.get("vat_amount") or 0),
            vat_rate=body.get("vat_rate") or "19%",
            date=body.get("date") or "",
            raw_text=body.get("raw_text") or "Manual entry",
            invoice_type=body.get("invoice_type") or "expense",
            invoice_number=body.get("invoice_number") or "",
            payment_method=body.get("payment_method") or "",
            category=body.get("category") or "other",
            processed=True,
        )
        db.add(inv)
        db.commit()
        db.refresh(inv)
        return {"success": True, "id": inv.id}
    except Exception:
        db.rollback()
        logger.exception("Create invoice failed")
        err(500, "Failed")
    finally:
        db.close()


@app.post("/invoices/upload-zip")
@limiter.limit("5/minute")
async def upload_zip(request: Request, file: UploadFile = File(...), invoice_type: str = "expense", user: dict = Depends(get_current_user)):
    """Upload a ZIP file containing invoices (PDF, JPG, PNG). Extracts and processes each file."""
    import zipfile
    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        err(400, "ZIP zu groß (max 50MB)")
    if not content[:4] == b"PK\x03\x04":
        err(400, "Keine gültige ZIP-Datei")
    results = []
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for name in zf.namelist():
                name_lower = name.lower()
                if name.startswith("__MACOSX") or name.startswith("."):
                    continue
                if not any(name_lower.endswith(ext) for ext in (".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".webp")):
                    results.append({"filename": name, "status": "skipped", "message": "Nicht unterstützt"})
                    continue
                try:
                    file_data = zf.read(name)
                    if len(file_data) == 0:
                        continue
                    if len(file_data) > MAX_FILE_SIZE:
                        results.append({"filename": name, "status": "error", "message": "Datei zu groß"})
                        continue
                    # Determine content type from extension
                    if name_lower.endswith(".pdf"):
                        ct = "application/pdf"
                    elif name_lower.endswith((".jpg", ".jpeg")):
                        ct = "image/jpeg"
                    elif name_lower.endswith(".png"):
                        ct = "image/png"
                    elif name_lower.endswith(".tiff"):
                        ct = "image/tiff"
                    else:
                        ct = "image/webp"
                    # Create a fake UploadFile for existing pipeline
                    fake_file = UploadFile(filename=name, file=io.BytesIO(file_data))
                    fake_file.content_type = ct
                    raw_text = ""
                    try:
                        raw_text = await asyncio.wait_for(extract_text(fake_file, handwriting=False, file_bytes=file_data), timeout=45)
                    except Exception:
                        logger.warning("OCR failed for ZIP entry: %s", name)
                    try:
                        parsed = parse_invoice(raw_text)
                    except Exception:
                        results.append({"filename": name, "status": "error", "message": "Parse failed"})
                        continue
                    if invoice_type in ("income", "expense"):
                        parsed["invoice_type"] = invoice_type
                    invoice_id = save_invoice(parsed, user_id=user["sub"], filename=name)
                    auto_create_cash_entry(invoice_id, user["sub"], parsed)
                    results.append({"filename": name, "status": "ok", "id": invoice_id, "vendor": parsed.get("vendor", ""), "total": parsed.get("total_amount", 0)})
                except Exception as e:
                    results.append({"filename": name, "status": "error", "message": str(e)})
    except zipfile.BadZipFile:
        err(400, "Beschädigte ZIP-Datei")
    return {"success": True, "results": results, "total": len(results)}


@app.post("/invoices/upload")
@limiter.limit("20/minute")
async def upload_invoice(request: Request, file: UploadFile = File(...), handwriting: bool = False, invoice_type: str = "expense", force_upload: bool = False, user: dict = Depends(get_current_user)):
    if file.content_type not in ALLOWED_TYPES:
        err(400, "Ungültige Datei. Erlaubt: PDF, JPG, PNG, TIFF, WEBP, ZIP")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        err(400, "Datei zu groß (max 10MB)")
    if len(content) == 0:
        err(400, "Leere Datei")

    # ZIP: extract and process each file inside
    if content[:4] == b"PK\x03\x04" or (file.content_type or "").lower() in ("application/zip", "application/x-zip-compressed"):
        import zipfile
        zip_results = []
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                for name in zf.namelist():
                    nl = name.lower()
                    if name.startswith("__MACOSX") or name.startswith("."):
                        continue
                    if not any(nl.endswith(e) for e in (".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".webp")):
                        continue
                    try:
                        fd = zf.read(name)
                        if not fd or len(fd) > MAX_FILE_SIZE:
                            continue
                        ct = "application/pdf" if nl.endswith(".pdf") else "image/jpeg" if nl.endswith((".jpg",".jpeg")) else "image/png"
                        fake = UploadFile(filename=name, file=io.BytesIO(fd))
                        fake.content_type = ct
                        rt = ""
                        try:
                            rt = await asyncio.wait_for(extract_text(fake, handwriting=handwriting, file_bytes=fd), timeout=45)
                        except Exception:
                            pass
                        parsed = parse_invoice(rt)
                        if invoice_type in ("income", "expense"):
                            parsed["invoice_type"] = invoice_type
                        inv_id = save_invoice(parsed, user_id=user["sub"], filename=name)
                        auto_create_cash_entry(inv_id, user["sub"], parsed)
                        zip_results.append({"filename": name, "status": "ok", "id": inv_id})
                    except Exception:
                        zip_results.append({"filename": name, "status": "error"})
        except zipfile.BadZipFile:
            err(400, "Beschädigte ZIP-Datei")
        return {"success": True, "results": zip_results, "count": len(zip_results)}

    if not _validate_file_magic(content, file.content_type or ""):
        err(400, "Ungültige Datei — Dateityp stimmt nicht mit Inhalt überein")

    await file.seek(0)

    # Save original file to DB for vault preview
    _file_data = content
    _file_ct = file.content_type or ""

    logger.info("Upload by user %s: type=%s, size=%d bytes", user["sub"], file.content_type, len(content))

    import gc
    raw_text = ""
    qr_data = {}
    try:
        raw_text, qr_data = await asyncio.wait_for(extract_text_and_qr(file, handwriting=handwriting, file_bytes=content), timeout=45)
    except asyncio.TimeoutError:
        logger.warning("OCR timeout — saving with empty text")
    except Exception:
        logger.warning("OCR failed — saving with empty text")
    finally:
        del content
        gc.collect()

    try:
        result = parse_invoice(raw_text)
    except Exception:
        logger.exception("Parsing failed for %s", file.filename)
        err(500, "Invoice parsing failed")

    # Merge QR data (QR overrides OCR if available)
    if qr_data:
        logger.info("QR data found: keys=%s", list(qr_data.keys()))
        if qr_data.get("company") and (not result.get("vendor") or result.get("vendor") == "Unbekannt"):
            result["vendor"] = qr_data["company"]
        if qr_data.get("amount") and (not result.get("total_amount") or result.get("total_amount") == 0):
            result["total_amount"] = qr_data["amount"]
        if qr_data.get("date") and (not result.get("date") or result["date"] == datetime.now().strftime("%Y-%m-%d")):
            result["date"] = qr_data["date"]
        if qr_data.get("invoice_number") and not result.get("invoice_number"):
            result["invoice_number"] = qr_data["invoice_number"]
        if qr_data.get("tax") and (not result.get("vat_amount") or result.get("vat_amount") == 0):
            result["vat_amount"] = qr_data["tax"]
        if qr_data.get("qr_raw"):
            result["raw_text"] = result.get("raw_text", "") + "\n\n[QR] " + qr_data["qr_raw"]

    # Duplicate check
    db_check = SessionLocal()
    try:
        dup = db_check.query(Invoice).filter(
            Invoice.user_id == user["sub"],
            Invoice.vendor == (result.get("vendor") or "Unbekannt"),
            Invoice.total_amount == safe_float(result.get("total_amount")),
            Invoice.date == (result.get("date") or ""),
        ).first()
        if dup:
            # --- ADDED START: force_upload bypass ---
            if not force_upload:
                # --- ADDED: check if duplicate is soft-deleted ---
                if dup.is_deleted:
                    logger.info("Duplicate is soft-deleted: invoice %d", dup.id)
                    return {"id": dup.id, "total_amount": safe_float(dup.total_amount), "filename": file.filename, "status": "duplicate_deleted", "duplicate_deleted": True, "message": "Bu fatura daha once silinmis. Geri yuklemek ister misiniz?"}
                # --- END ---
                logger.info("Duplicate detected: vendor=%s, amount=%s, date=%s", result.get("vendor"), result.get("total_amount"), result.get("date"))
                return {"id": dup.id, "total_amount": safe_float(dup.total_amount), "filename": file.filename, "status": "duplicate", "duplicate": True, "can_force": True, "message": "Bu fatura zaten yuklu. Tekrar yuklemek ister misiniz?"}
            else:
                logger.info("Force upload: duplicate bypassed for vendor=%s, amount=%s", result.get("vendor"), result.get("total_amount"))
            # --- ADDED END ---
    finally:
        db_check.close()

    if invoice_type in ("income", "expense"):
        result["invoice_type"] = invoice_type

    try:
        invoice_id = save_invoice(result, user_id=user["sub"], filename=file.filename, file_data=_file_data, file_content_type=_file_ct)
    except Exception:
        logger.exception("DB save failed")
        err(500, "Failed to save invoice")

    auto_create_cash_entry(invoice_id, user["sub"], result)

    # Auto-detect income: if vendor/IBAN/email matches user's registered company
    try:
        db_c = SessionLocal()
        user_companies = db_c.query(UserCompany).filter(UserCompany.user_id == user["sub"]).all()
        if user_companies:
            inv = db_c.query(Invoice).filter(Invoice.id == invoice_id).first()
            if inv:
                vendor_lower = (inv.vendor or "").lower()
                inv_iban = result.get("vendor_iban", "").replace(" ", "").upper()
                inv_email = result.get("vendor_email", "").lower()
                for uc in user_companies:
                    matched = False
                    # Match by company name
                    if vendor_lower and uc.company_name:
                        if uc.company_name.lower() in vendor_lower or vendor_lower in uc.company_name.lower() or _fuzzy_match(uc.company_name, inv.vendor or ""):
                            matched = True
                    # Match by IBAN
                    if not matched and inv_iban and uc.iban:
                        if inv_iban == uc.iban.replace(" ", "").upper():
                            matched = True
                    # Match by email
                    if not matched and inv_email and uc.email:
                        if inv_email == uc.email.lower():
                            matched = True
                    if matched:
                        inv.invoice_type = "income"
                        logger.info("Auto-detected income: invoice %d matches company '%s'", invoice_id, uc.company_name)
                        db_c.commit()
                        break
        db_c.close()
    except Exception:
        pass

    # OCR quality warning
    _ocr_warning = ""
    if not raw_text or len(raw_text.strip()) < 20:
        _ocr_warning = "OCR konnte den Text nicht lesen — bitte manuell prüfen"
    elif safe_float(result.get("total_amount")) == 0:
        _ocr_warning = "Betrag nicht erkannt — bitte manuell eingeben"
    elif result.get("vendor") == "Unbekannt":
        _ocr_warning = "Lieferant nicht erkannt — bitte manuell eingeben"

    return {
        "id": invoice_id,
        "total_amount": safe_float(result.get("total_amount")),
        "filename": file.filename,
        "status": "ok",
        "warning": _ocr_warning,
        "vendor": result.get("vendor", ""),
        "vendor_iban": result.get("vendor_iban", ""),
        "vendor_email": result.get("vendor_email", ""),
        "vendor_phone": result.get("vendor_phone", ""),
        "vendor_address": result.get("vendor_address", ""),
    }


@app.post("/invoices/batch")
async def upload_batch(files: List[UploadFile] = File(...), invoice_type: str = "expense", user: dict = Depends(get_current_user)):
    import gc
    results = []
    for file in files:
        try:
            if file.content_type not in ALLOWED_TYPES:
                results.append({"filename": file.filename, "status": "error", "message": "Ungültige Datei"})
                continue
            content = await file.read()
            if len(content) > MAX_FILE_SIZE:
                results.append({"filename": file.filename, "status": "error", "message": "Datei zu groß"})
                continue
            if len(content) == 0:
                results.append({"filename": file.filename, "status": "error", "message": "Leere Datei"})
                continue
            if not _validate_file_magic(content, file.content_type or ""):
                results.append({"filename": file.filename, "status": "error", "message": "Ungültige Datei"})
                continue
            await file.seek(0)
            raw_text = ""
            try:
                raw_text = await asyncio.wait_for(extract_text(file, handwriting=False, file_bytes=content), timeout=45)
            except Exception:
                logger.warning("OCR failed/timeout for batch file")
            finally:
                del content
                gc.collect()
            try:
                parsed = parse_invoice(raw_text)
            except Exception:
                results.append({"filename": file.filename, "status": "error", "message": "Parse failed"})
                continue
            # Duplicate check
            db_dup = SessionLocal()
            try:
                dup = db_dup.query(Invoice).filter(
                    Invoice.user_id == user["sub"],
                    Invoice.vendor == (parsed.get("vendor") or "Unbekannt"),
                    Invoice.total_amount == safe_float(parsed.get("total_amount")),
                    Invoice.date == (parsed.get("date") or ""),
                ).first()
            finally:
                db_dup.close()
            if dup:
                results.append({"filename": file.filename, "status": "duplicate", "message": "Duplikat erkannt"})
                continue
            if invoice_type in ("income", "expense"):
                parsed["invoice_type"] = invoice_type
            # --- ADDED: re-read file for storage (content was deleted after OCR) ---
            await file.seek(0)
            _batch_file_data = await file.read()
            invoice_id = save_invoice(parsed, user_id=user["sub"], filename=file.filename, file_data=_batch_file_data, file_content_type=file.content_type or "")
            auto_create_cash_entry(invoice_id, user["sub"], parsed)
            results.append({
                "filename": file.filename,
                "status": "ok",
                "message": f"OK — €{safe_float(parsed.get('total_amount')):.2f}",
                "id": invoice_id,
            })
        except Exception as e:
            results.append({"filename": file.filename, "status": "error", "message": str(e)})
    return {"results": results}


# ============================================================
# INVOICES: LIST
# ============================================================

@app.get("/invoices")
def list_invoices(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    search: Optional[str] = Query(None),
    vendor: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
):
    db = SessionLocal()
    try:
        q = db.query(Invoice).filter(Invoice.user_id == user["sub"])
        # --- ADDED: exclude soft-deleted ---
        q = q.filter((Invoice.is_deleted == False) | (Invoice.is_deleted == None))
        # --- END ---

        if search:
            # Smart multi-keyword search: split by space, ALL keywords must match
            # Search across vendor, category, and raw_text (OCR content)
            import unicodedata
            normalized = unicodedata.normalize("NFKD", search.lower().strip())
            keywords = [k.strip() for k in normalized.split() if k.strip()]
            from sqlalchemy import or_
            for kw in keywords:
                pattern = f"%{kw}%"
                q = q.filter(or_(
                    Invoice.raw_text.ilike(pattern),
                    Invoice.vendor.ilike(pattern),
                    Invoice.category.ilike(pattern),
                    Invoice.invoice_number.ilike(pattern),
                ))
        if vendor:
            q = q.filter(Invoice.vendor.ilike(f"%{vendor}%"))
        if status == "processed":
            q = q.filter(Invoice.processed == True)
        elif status == "unprocessed":
            q = q.filter(Invoice.processed == False)
        if category:
            q = q.filter(Invoice.category == category)
        # Validate date range (reject invalid years like 333333)
        import re as _re
        _current_year = datetime.now().year
        if date_from and _re.match(r"^\d{4}-\d{2}-\d{2}$", date_from):
            if 2020 <= int(date_from[:4]) <= _current_year + 1:
                q = q.filter(Invoice.date >= date_from)
        if date_to and _re.match(r"^\d{4}-\d{2}-\d{2}$", date_to):
            if 2020 <= int(date_to[:4]) <= _current_year + 1:
                q = q.filter(Invoice.date <= date_to)

        total_count = q.count()
        q = q.order_by(Invoice.created_at.desc())
        invoices = q.offset(skip).limit(limit).all()

        return ok_list(
            [invoice_to_dict(i) for i in invoices],
            total_count,
        )
    except HTTPException:
        raise
    except Exception:
        logger.exception("Failed to list invoices")
        err(500, "Failed to load invoices")
    finally:
        db.close()


# ============================================================
# INVOICES: DASHBOARD
# ============================================================

@app.get("/invoices/dashboard")
def invoice_dashboard(country: str = Query("DE"), user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        all_invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"], (Invoice.is_deleted == False) | (Invoice.is_deleted == None)).all()
        # Filter out invalid entries (amount=0 or vendor=Unbekannt)
        invoices = [i for i in all_invoices if safe_float(i.total_amount) > 0 and safe_vendor(i.vendor) != "Unbekannt"]

        # Use ONLY invoices as source of truth (no cash_entries to avoid double counting)
        inv_inc = [i for i in invoices if safe_invoice_type(i.invoice_type) == "income"]
        inv_exp = [i for i in invoices if safe_invoice_type(i.invoice_type) == "expense"]

        total_income = sum(safe_float(i.total_amount) for i in inv_inc)
        total_expenses = sum(safe_float(i.total_amount) for i in inv_exp)
        net_profit = total_income - total_expenses

        total_vat_paid = sum(safe_float(i.vat_amount) for i in inv_exp)
        total_vat_collected = sum(safe_float(i.vat_amount) for i in inv_inc)
        vat_balance = total_vat_collected - total_vat_paid

        if country == "DE":
            if net_profit > 277826:
                tax_rate = 45
            elif net_profit > 61356:
                tax_rate = 42
            elif net_profit > 17005:
                tax_rate = 30
            elif net_profit > 10908:
                tax_rate = 14
            else:
                tax_rate = 0
        else:
            tax_rate = 30

        tax_estimate = round(net_profit * tax_rate / 100, 2) if net_profit > 0 else 0

        month_map = {}
        for i in invoices:
            d = safe_date_str(i.date)
            if not d or len(d) < 7 or "-" not in d:
                continue
            m = d[:7]
            if m not in month_map:
                month_map[m] = {"month": m, "income": 0.0, "expenses": 0.0}
            if safe_invoice_type(i.invoice_type) == "income":
                month_map[m]["income"] += safe_float(i.total_amount)
            else:
                month_map[m]["expenses"] += safe_float(i.total_amount)
        monthly_breakdown = sorted(month_map.values(), key=lambda x: x["month"])
        for mb in monthly_breakdown:
            mb["income"] = round(mb["income"], 2)
            mb["expenses"] = round(mb["expenses"], 2)

        cat_map = {}
        for i in invoices:
            c = safe_category(i.category)
            cat_map[c] = cat_map.get(c, 0) + safe_float(i.total_amount)
        by_category = [{"category": k, "total": round(v, 2)} for k, v in sorted(cat_map.items(), key=lambda x: -x[1])]

        return {
            "total_income": round(total_income, 2),
            "total_expenses": round(total_expenses, 2),
            "net_profit": round(net_profit, 2),
            "tax_estimate": tax_estimate,
            "tax_rate_applied": tax_rate,
            "income_count": len(inv_inc),
            "expense_count": len(inv_exp),
            "invoice_count": len(invoices),
            "invalid_count": len(all_invoices) - len(invoices),
            "monthly_breakdown": monthly_breakdown,
            "by_category": by_category,
            "total_vat_paid": round(total_vat_paid, 2),
            "total_vat_collected": round(total_vat_collected, 2),
            "vat_balance": round(vat_balance, 2),
        }
    except Exception:
        logger.exception("Dashboard failed")
        err(500, "Dashboard failed")
    finally:
        db.close()


# ============================================================
# INVOICES: SUMMARY
# ============================================================

@app.get("/invoices/summary")
def invoice_summary(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        all_invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"]).all()
        invoices = [i for i in all_invoices if safe_float(i.total_amount) > 0 and safe_vendor(i.vendor) != "Unbekannt"]
        total_count = len(invoices)
        processed = sum(1 for i in invoices if i.processed)
        unprocessed = total_count - processed
        total_revenue = sum(safe_float(i.total_amount) for i in invoices)
        return {
            "success": True,
            "total_count": total_count,
            "processed": processed,
            "unprocessed": unprocessed,
            "total_revenue": round(total_revenue, 2),
        }
    except Exception:
        logger.exception("Summary failed")
        err(500, "Failed to load summary")
    finally:
        db.close()


# ============================================================
# INVOICES: UPDATE (PATCH + PUT)
# ============================================================

class InvoiceUpdate(BaseModel):
    vendor: Optional[str] = None
    category: Optional[str] = None
    total_amount: Optional[float] = None
    vat_amount: Optional[float] = None
    vat_rate: Optional[str] = None
    date: Optional[str] = None
    invoice_type: Optional[str] = None
    invoice_number: Optional[str] = None
    payment_method: Optional[str] = None
    processed: Optional[bool] = None


def _do_update_invoice(invoice_id: int, body: InvoiceUpdate, user: dict):
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        if body.vendor is not None:
            inv.vendor = body.vendor
        if body.category is not None:
            inv.category = body.category
        if body.total_amount is not None:
            inv.total_amount = body.total_amount
        if body.vat_amount is not None:
            inv.vat_amount = body.vat_amount
        if body.vat_rate is not None:
            inv.vat_rate = body.vat_rate
        if body.date is not None:
            inv.date = body.date
        if body.invoice_type is not None:
            inv.invoice_type = body.invoice_type
        if body.invoice_number is not None:
            inv.invoice_number = body.invoice_number
        if body.payment_method is not None:
            inv.payment_method = body.payment_method
        if body.processed is not None:
            inv.processed = body.processed
        db.commit()
        db.refresh(inv)
        return {"success": True, **invoice_to_dict(inv)}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Update invoice failed")
        err(500, "Failed to update invoice")
    finally:
        db.close()


@app.patch("/invoices/{invoice_id}")
def patch_invoice(invoice_id: int, body: InvoiceUpdate, user: dict = Depends(get_current_user)):
    return _do_update_invoice(invoice_id, body, user)


@app.put("/invoices/{invoice_id}")
def put_invoice(invoice_id: int, body: InvoiceUpdate, user: dict = Depends(get_current_user)):
    return _do_update_invoice(invoice_id, body, user)


# --- ADDED START: Single invoice detail with full OCR text ---
@app.get("/invoices/{invoice_id}/detail")
def get_invoice_detail(invoice_id: int, user: dict = Depends(get_current_user)):
    """Get full invoice detail including complete raw OCR text."""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        result = invoice_to_dict(inv)
        result["raw_text"] = inv.raw_text or ""
        return result
    finally:
        db.close()
# --- ADDED END ---

# ============================================================
# INVOICES: DELETE
# ============================================================

@app.delete("/invoices/{invoice_id}")
def delete_invoice(invoice_id: int, permanent: bool = False, user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        # --- ADDED: soft delete ---
        if permanent:
            db.delete(inv)
            logger.info("Permanent delete: invoice %d", invoice_id)
        else:
            inv.is_deleted = True
            inv.deleted_at = datetime.now()
            logger.info("Soft delete: invoice %d", invoice_id)
            # Also soft-delete linked cash entry
            linked = db.query(CashEntry).filter(CashEntry.invoice_id == invoice_id, CashEntry.user_id == user["sub"]).first()
            if linked:
                linked.is_deleted = True
                linked.deleted_at = datetime.now()
        # --- END ---
        db.commit()
        return {"success": True, "deleted": invoice_id}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Delete invoice failed")
        err(500, "Failed to delete invoice")
    finally:
        db.close()


class BulkDeleteRequest(BaseModel):
    ids: List[int]


@app.post("/invoices/bulk-delete")
def bulk_delete_invoices(body: BulkDeleteRequest, user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        # --- ADDED: soft delete ---
        invs = db.query(Invoice).filter(Invoice.id.in_(body.ids), Invoice.user_id == user["sub"]).all()
        deleted = 0
        for inv in invs:
            inv.is_deleted = True
            inv.deleted_at = datetime.now()
            deleted += 1
            linked = db.query(CashEntry).filter(CashEntry.invoice_id == inv.id, CashEntry.user_id == user["sub"]).first()
            if linked:
                linked.is_deleted = True
                linked.deleted_at = datetime.now()
        logger.info("Soft bulk delete: %d invoices", deleted)
        # --- END ---
        db.commit()
        return {"success": True, "deleted": deleted}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Bulk delete failed")
        err(500, "Bulk delete failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: MODELS
# ============================================================

class CashEntryCreate(BaseModel):
    description: str
    gross_amount: float
    entry_type: str
    vendor: Optional[str] = None
    category: Optional[str] = None
    vat_rate: Optional[str] = None
    payment_method: Optional[str] = None
    reference: Optional[str] = None
    notes: Optional[str] = None
    date: Optional[str] = None


class CashEntryUpdate(BaseModel):
    description: Optional[str] = None
    gross_amount: Optional[float] = None
    entry_type: Optional[str] = None
    vendor: Optional[str] = None
    category: Optional[str] = None
    vat_rate: Optional[str] = None
    payment_method: Optional[str] = None
    reference: Optional[str] = None
    notes: Optional[str] = None
    date: Optional[str] = None


# ============================================================
# BOOKKEEPING: LIST (GET /bookkeeping + /kassenbuch)
# ============================================================

def _list_bookkeeping(skip, limit, user):
    db = SessionLocal()
    try:
        q = db.query(CashEntry).filter(CashEntry.user_id == user["sub"])
        # --- ADDED: exclude soft-deleted ---
        q = q.filter((CashEntry.is_deleted == False) | (CashEntry.is_deleted == None))
        # --- END ---
        total_count = q.count()
        all_entries = q.all()
        entries = q.order_by(CashEntry.date.desc()).offset(skip).limit(limit).all()
        # Calculate totals across ALL entries (not just current page)
        total_gross = sum(safe_float(e.gross_amount) for e in all_entries)
        total_vat = sum(safe_float(e.vat_amount) for e in all_entries)
        total_income = sum(safe_float(e.gross_amount) for e in all_entries if e.entry_type == "income")
        total_expense = sum(safe_float(e.gross_amount) for e in all_entries if e.entry_type == "expense")
        return {
            "success": True,
            "items": [cash_entry_to_dict(e) for e in entries],
            "total": total_count,
            "summary": {
                "total_gross": round(total_gross, 2),
                "total_vat": round(total_vat, 2),
                "total_income": round(total_income, 2),
                "total_expense": round(total_expense, 2),
                "total_expenses": round(total_expense, 2),
                "net": round(total_income - total_expense, 2),
                "net_profit": round(total_income - total_expense, 2),
                "vat_balance": round(
                    sum(safe_float(e.vat_amount) for e in all_entries if e.entry_type == "income") -
                    sum(safe_float(e.vat_amount) for e in all_entries if e.entry_type == "expense"), 2),
                "entry_count": total_count,
            },
        }
    except Exception:
        logger.exception("Failed to list cash entries")
        err(500, "Failed to load cash entries")
    finally:
        db.close()


@app.get("/bookkeeping")
def list_bookkeeping(skip: int = Query(0, ge=0), limit: int = Query(50, ge=1, le=200), user: dict = Depends(get_current_user)):
    return _list_bookkeeping(skip, limit, user)


@app.get("/kassenbuch")
def list_kassenbuch(skip: int = Query(0, ge=0), limit: int = Query(50, ge=1, le=200), user: dict = Depends(get_current_user)):
    return _list_bookkeeping(skip, limit, user)


# ============================================================
# BOOKKEEPING: CREATE (POST /bookkeeping + /kassenbuch)
# ============================================================

def _create_bookkeeping(body: CashEntryCreate, user: dict):
    if body.entry_type not in ("income", "expense"):
        err(400, "entry_type must be 'income' or 'expense'")
    db = SessionLocal()
    try:
        entry_date = parse_date_str_to_datetime(body.date)
        vat_amount = calc_vat(body.gross_amount, body.vat_rate)
        entry = CashEntry(
            user_id=user["sub"],
            description=body.description,
            gross_amount=body.gross_amount,
            vat_amount=vat_amount,
            vat_rate=body.vat_rate or "0%",
            vendor=body.vendor or "Unbekannt",
            entry_type=body.entry_type,
            category=body.category or "other",
            payment_method=body.payment_method or "",
            reference=body.reference or "",
            notes=body.notes or "",
            date=entry_date,
        )
        db.add(entry)
        db.commit()
        db.refresh(entry)
        # Also create a corresponding Invoice so it appears in dashboard
        try:
            inv = Invoice(
                user_id=user["sub"],
                filename=None,
                vendor=body.vendor or "Manual Entry",
                total_amount=body.gross_amount or 0.0,
                vat_amount=vat_amount,
                vat_rate=body.vat_rate or "0%",
                date=body.date or "",
                raw_text=f"manual entry: {body.description}",
                invoice_type=body.entry_type,
                invoice_number="",
                payment_method=body.payment_method or "",
                category=body.category or "other",
                processed=True,
            )
            db.add(inv)
            db.commit()
            logger.info("Auto-created invoice from manual Kassenbuch entry %s", entry.id)
        except Exception:
            logger.exception("Failed to auto-create invoice from Kassenbuch entry")
        return {"success": True, **cash_entry_to_dict(entry)}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Create cash entry failed")
        err(500, "Failed to create entry")
    finally:
        db.close()


@app.post("/bookkeeping")
def create_bookkeeping(body: CashEntryCreate, user: dict = Depends(get_current_user)):
    return _create_bookkeeping(body, user)


@app.post("/kassenbuch")
def create_kassenbuch(body: CashEntryCreate, user: dict = Depends(get_current_user)):
    return _create_bookkeeping(body, user)


# ============================================================
# BOOKKEEPING: UPDATE (PATCH+PUT /bookkeeping/{id} + /kassenbuch/{id})
# ============================================================

def _update_bookkeeping(entry_id: int, body: CashEntryUpdate, user: dict):
    db = SessionLocal()
    try:
        entry = db.query(CashEntry).filter(CashEntry.id == entry_id, CashEntry.user_id == user["sub"]).first()
        if not entry:
            err(404, "Entry not found")
        if body.description is not None:
            entry.description = body.description
        if body.gross_amount is not None:
            entry.gross_amount = body.gross_amount
        if body.entry_type is not None:
            if body.entry_type not in ("income", "expense"):
                err(400, "entry_type must be 'income' or 'expense'")
            entry.entry_type = body.entry_type
        if body.vendor is not None:
            entry.vendor = body.vendor
        if body.category is not None:
            entry.category = body.category
        if body.vat_rate is not None:
            entry.vat_rate = body.vat_rate
        if body.payment_method is not None:
            entry.payment_method = body.payment_method
        if body.reference is not None:
            entry.reference = body.reference
        if body.notes is not None:
            entry.notes = body.notes
        if body.date is not None:
            entry.date = parse_date_str_to_datetime(body.date)
        if body.gross_amount is not None or body.vat_rate is not None:
            entry.vat_amount = calc_vat(entry.gross_amount, entry.vat_rate)
        db.commit()
        db.refresh(entry)
        return {"success": True, **cash_entry_to_dict(entry)}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Update cash entry failed")
        err(500, "Failed to update entry")
    finally:
        db.close()


@app.patch("/bookkeeping/{entry_id}")
def patch_bookkeeping(entry_id: int, body: CashEntryUpdate, user: dict = Depends(get_current_user)):
    return _update_bookkeeping(entry_id, body, user)


@app.put("/bookkeeping/{entry_id}")
def put_bookkeeping(entry_id: int, body: CashEntryUpdate, user: dict = Depends(get_current_user)):
    return _update_bookkeeping(entry_id, body, user)


@app.patch("/kassenbuch/{entry_id}")
def patch_kassenbuch(entry_id: int, body: CashEntryUpdate, user: dict = Depends(get_current_user)):
    return _update_bookkeeping(entry_id, body, user)


@app.put("/kassenbuch/{entry_id}")
def put_kassenbuch(entry_id: int, body: CashEntryUpdate, user: dict = Depends(get_current_user)):
    return _update_bookkeeping(entry_id, body, user)


# ============================================================
# BOOKKEEPING: DELETE (/bookkeeping/{id} + /kassenbuch/{id})
# ============================================================

def _delete_bookkeeping(entry_id: int, user: dict):
    db = SessionLocal()
    try:
        entry = db.query(CashEntry).filter(CashEntry.id == entry_id, CashEntry.user_id == user["sub"]).first()
        if not entry:
            err(404, "Entry not found")
        # --- ADDED: soft delete ---
        entry.is_deleted = True
        entry.deleted_at = datetime.now()
        logger.info("Soft delete: cash entry %d", entry_id)
        # --- END ---
        db.commit()
        return {"success": True, "deleted": entry_id}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Delete cash entry failed")
        err(500, "Failed to delete entry")
    finally:
        db.close()


@app.delete("/bookkeeping/{entry_id}")
def delete_bookkeeping(entry_id: int, user: dict = Depends(get_current_user)):
    return _delete_bookkeeping(entry_id, user)


@app.delete("/kassenbuch/{entry_id}")
def delete_kassenbuch(entry_id: int, user: dict = Depends(get_current_user)):
    return _delete_bookkeeping(entry_id, user)


# ============================================================
# BOOKKEEPING: SYNC INVOICES
# ============================================================

@app.post("/bookkeeping/sync-invoices")
def sync_invoices_to_bookkeeping(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"]).all()
        existing_invoice_ids = set()
        all_entries = db.query(CashEntry).filter(CashEntry.user_id == user["sub"]).all()
        for e in all_entries:
            if e.invoice_id:
                existing_invoice_ids.add(e.invoice_id)

        synced = 0
        skipped = 0
        for inv in invoices:
            if inv.id in existing_invoice_ids:
                skipped += 1
                continue
            vat_amount = calc_vat(safe_float(inv.total_amount), safe_vat_rate(inv.vat_rate))
            entry_date = parse_date_str_to_datetime(inv.date) if inv.date else inv.created_at
            entry = CashEntry(
                user_id=user["sub"],
                description=safe_vendor(inv.vendor),
                vendor=safe_vendor(inv.vendor),
                gross_amount=safe_float(inv.total_amount),
                vat_amount=vat_amount,
                vat_rate=safe_vat_rate(inv.vat_rate),
                entry_type=safe_invoice_type(inv.invoice_type),
                category=safe_category(inv.category),
                payment_method=safe_str(inv.payment_method),
                invoice_id=inv.id,
                date=entry_date,
            )
            db.add(entry)
            synced += 1
        # Reverse sync: CashEntry → Invoice (for manual entries without invoice)
        existing_inv_vendors_amounts = set()
        for inv in invoices:
            existing_inv_vendors_amounts.add(f"{(inv.vendor or '').lower()}-{safe_float(inv.total_amount)}")
        rev_synced = 0
        for entry in all_entries:
            if entry.invoice_id:
                continue  # already linked to an invoice
            # Duplicate check: vendor + amount
            dup_key = f"{(entry.vendor or entry.description or '').lower()}-{safe_float(entry.gross_amount)}"
            if dup_key in existing_inv_vendors_amounts:
                continue  # already has matching invoice
            inv = Invoice(
                user_id=user["sub"],
                filename=None,
                vendor=entry.vendor or "Manual Entry",
                total_amount=safe_float(entry.gross_amount),
                vat_amount=safe_float(entry.vat_amount),
                vat_rate=entry.vat_rate or "0%",
                date=entry.date.strftime("%Y-%m-%d") if entry.date else "",
                raw_text=f"Sync from Kassenbuch: {safe_str(entry.description)}",
                invoice_type=entry.entry_type or "expense",
                invoice_number="",
                payment_method=safe_str(entry.payment_method),
                category=safe_category(entry.category),
                processed=True,
            )
            db.add(inv)
            rev_synced += 1

        db.commit()
        return {"synced": synced, "skipped": skipped, "reverse_synced": rev_synced}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Sync invoices failed")
        err(500, "Sync failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: RECONCILE
# ============================================================

@app.post("/bookkeeping/{entry_id}/reconcile")
def reconcile_entry(entry_id: int, user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        entry = db.query(CashEntry).filter(CashEntry.id == entry_id, CashEntry.user_id == user["sub"]).first()
        if not entry:
            err(404, "Entry not found")
        entry.is_reconciled = not entry.is_reconciled
        db.commit()
        db.refresh(entry)
        return {"success": True, **cash_entry_to_dict(entry)}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Reconcile failed")
        err(500, "Reconcile failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: SUMMARY
# ============================================================

@app.get("/bookkeeping/summary/overview")
def bookkeeping_summary(year: int = Query(None), user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        q = db.query(CashEntry).filter(CashEntry.user_id == user["sub"])
        if year:
            q = q.filter(CashEntry.date >= datetime(year, 1, 1))
            q = q.filter(CashEntry.date < datetime(year + 1, 1, 1))
        entries = q.all()
        total_income = sum(safe_float(e.gross_amount) for e in entries if e.entry_type == "income")
        total_expenses = sum(safe_float(e.gross_amount) for e in entries if e.entry_type == "expense")
        vat_collected = sum(safe_float(e.vat_amount) for e in entries if e.entry_type == "income")
        vat_paid = sum(safe_float(e.vat_amount) for e in entries if e.entry_type == "expense")
        return {
            "total_income": round(total_income, 2),
            "total_expenses": round(total_expenses, 2),
            "net_profit": round(total_income - total_expenses, 2),
            "vat_balance": round(vat_collected - vat_paid, 2),
            "entry_count": len(entries),
        }
    except Exception:
        logger.exception("Summary failed")
        err(500, "Summary failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: EXPORT CSV
# ============================================================

@app.get("/bookkeeping/export/csv")
def export_bookkeeping_csv(year: int = Query(None), user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        q = db.query(CashEntry).filter(CashEntry.user_id == user["sub"])
        if year:
            q = q.filter(CashEntry.date >= datetime(year, 1, 1))
            q = q.filter(CashEntry.date < datetime(year + 1, 1, 1))
        entries = q.order_by(CashEntry.date.desc()).all()
        buf = io.StringIO()
        buf.write("Datum,Typ,Beschreibung,Lieferant,Betrag,MwSt,MwSt-Satz,Kategorie,Zahlungsart,Beleg-Nr.\n")
        for e in entries:
            date_str = e.date.strftime("%d.%m.%Y") if e.date else ""
            desc = (e.description or "").replace('"', '""')
            vendor = (e.vendor or "").replace('"', '""')
            buf.write(f'{date_str},{e.entry_type or ""},"{desc}","{vendor}",{safe_float(e.gross_amount):.2f},{safe_float(e.vat_amount):.2f},{safe_vat_rate(e.vat_rate)},{safe_category(e.category)},{safe_str(e.payment_method)},{safe_str(e.reference)}\n')
        buf.seek(0)
        return StreamingResponse(buf, media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=kassenbuch_{year or 'all'}.csv"})
    except Exception:
        logger.exception("Bookkeeping CSV export failed")
        err(500, "Export failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: CSV IMPORT
# ============================================================

@app.post("/bookkeeping/import-csv")
async def import_csv(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    import csv
    content = await file.read()
    text = content.decode("utf-8", errors="ignore")
    first_line = text.split("\n")[0] if text else ""
    delimiter = ";" if ";" in first_line else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

    db = SessionLocal()
    imported = 0
    errors = []
    try:
        for idx, row in enumerate(reader, 1):
            try:
                # Flexible column mapping — supports Export format + custom formats
                vendor = row.get("Lieferant") or row.get("Vendor") or row.get("vendor") or ""
                beschreibung = row.get("Beschreibung") or row.get("description") or row.get("Description") or vendor or ""
                datum = row.get("Datum") or row.get("date") or row.get("Date") or ""
                betrag_raw = row.get("Betrag") or row.get("Ausgaben") or row.get("amount") or row.get("expenses") or "0"
                einnahmen = row.get("Einnahmen") or row.get("income") or "0"
                typ = row.get("Typ") or row.get("type") or row.get("Type") or ""
                category = row.get("Kategorie") or row.get("category") or row.get("Category") or "other"
                payment = row.get("Zahlungsart") or row.get("Zahlungsmethode") or row.get("payment_method") or ""
                mwst_raw = row.get("MwSt") or row.get("vat_amount") or ""
                mwst_satz = row.get("MwSt-Satz") or row.get("vat_rate") or "19%"
                inv_nr = row.get("Rechnungs-Nr.") or row.get("invoice_number") or ""
                if not vendor:
                    vendor = beschreibung[:50] or "Import"

                def _parse_num(s):
                    return float(str(s).replace(",", ".").replace("€", "").replace("%", "").replace(" ", "").strip() or "0")

                betrag_val = _parse_num(betrag_raw)
                einnahmen_val = _parse_num(einnahmen)

                # Determine type: from Typ column, or from Einnahmen column
                if typ.lower() in ("income", "einnahme", "einnahmen"):
                    entry_type = "income"
                    amount = betrag_val if betrag_val > 0 else einnahmen_val
                elif typ.lower() in ("expense", "ausgabe", "ausgaben"):
                    entry_type = "expense"
                    amount = betrag_val
                elif einnahmen_val > 0:
                    entry_type = "income"
                    amount = einnahmen_val
                else:
                    entry_type = "expense"
                    amount = betrag_val

                if amount <= 0 and not beschreibung:
                    continue

                # Duplicate check
                existing = db.query(CashEntry).filter(
                    CashEntry.user_id == user["sub"],
                    CashEntry.description == (beschreibung or vendor),
                    CashEntry.gross_amount == amount,
                ).first()
                if existing:
                    continue

                date_val = None
                if datum:
                    parts = datum.strip().split(".")
                    if len(parts) == 3:
                        try:
                            date_val = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
                        except ValueError:
                            pass
                    if not date_val:
                        date_val = parse_date_str_to_datetime(datum)
                if not date_val:
                    date_val = datetime.now()

                # Use MwSt from CSV if provided, otherwise calculate
                if mwst_raw:
                    vat_amount = _parse_num(mwst_raw)
                else:
                    rate = _parse_num(mwst_satz) if mwst_satz else 19
                    vat_amount = round(amount * rate / (100 + rate), 2) if amount > 0 else 0
                vat_rate_str = mwst_satz if mwst_satz else "19%"
                if "%" not in vat_rate_str:
                    vat_rate_str += "%"

                entry = CashEntry(
                    user_id=user["sub"],
                    description=beschreibung or vendor,
                    vendor=vendor,
                    gross_amount=amount,
                    vat_amount=vat_amount,
                    vat_rate=vat_rate_str,
                    entry_type=entry_type,
                    category=category,
                    payment_method=payment,
                    reference=f"CSV-Import Zeile {idx}",
                    notes="Importiert aus CSV",
                    date=date_val,
                )
                db.add(entry)

                inv = Invoice(
                    user_id=user["sub"],
                    filename=f"csv-import-{idx}",
                    vendor=vendor,
                    total_amount=amount,
                    vat_amount=vat_amount,
                    vat_rate=vat_rate_str,
                    date=date_val.strftime("%Y-%m-%d") if date_val else "",
                    raw_text=f"CSV Import: {beschreibung}",
                    invoice_type=entry_type,
                    invoice_number=inv_nr,
                    payment_method=payment,
                    category=category,
                    processed=True,
                )
                db.add(inv)
                imported += 1
            except Exception as e:
                errors.append(f"Zeile {idx}: {str(e)[:80]}")
        db.commit()
        return {"success": True, "imported": imported, "errors": errors}
    except Exception:
        db.rollback()
        logger.exception("CSV import failed")
        err(500, "CSV import failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: XLSX IMPORT
# ============================================================

@app.post("/bookkeeping/import-xlsx")
async def import_xlsx(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Import Excel (.xlsx) file into Kassenbuch + Rechnungen."""
    from openpyxl import load_workbook
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    db = SessionLocal()
    imported = 0
    errors = []
    try:
        headers = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1:
                headers = [str(c or "").strip().lower() for c in row]
                continue
            if not any(row):
                continue
            rd = dict(zip(headers, [c for c in row]))

            def _col(names):
                for n in names:
                    v = rd.get(n) or rd.get(n.lower())
                    if v is not None and str(v).strip():
                        return str(v).strip()
                return ""

            beschreibung = _col(["beschreibung", "description", "lieferant", "vendor"])
            datum = _col(["datum", "date"])
            vendor = _col(["lieferant", "vendor"]) or beschreibung[:50] or "Import"
            category = _col(["kategorie", "category"]) or "other"
            payment = _col(["zahlungsart", "zahlungsmethode", "payment_method"]) or ""
            inv_nr = _col(["rechnungs-nr.", "invoice_number"]) or ""
            typ = _col(["typ", "type"]) or ""

            def _num(names):
                raw = _col(names)
                if not raw:
                    return 0.0
                return float(str(raw).replace(",", ".").replace("€", "").replace(" ", "").strip() or "0")

            betrag = _num(["betrag", "ausgaben", "amount", "expenses"])
            einnahmen = _num(["einnahmen", "income"])
            mwst = _num(["mwst", "vat_amount"])
            mwst_satz = _col(["mwst-satz", "vat_rate"]) or "19%"
            if "%" not in mwst_satz:
                mwst_satz += "%"

            if typ.lower() in ("income", "einnahme", "einnahmen"):
                entry_type = "income"
                amount = betrag if betrag > 0 else einnahmen
            elif einnahmen > 0:
                entry_type = "income"
                amount = einnahmen
            else:
                entry_type = "expense"
                amount = betrag

            if amount <= 0 and not beschreibung:
                continue

            date_val = parse_date_str_to_datetime(str(datum)) if datum else None
            if not date_val:
                date_val = datetime.now()
            if not mwst and amount > 0:
                rate = float(mwst_satz.replace("%", "").replace(",", ".").strip() or "19")
                mwst = round(amount * rate / (100 + rate), 2)

            try:
                entry = CashEntry(user_id=user["sub"], description=beschreibung or vendor, vendor=vendor,
                    gross_amount=amount, vat_amount=mwst, vat_rate=mwst_satz, entry_type=entry_type,
                    category=category, payment_method=payment, reference=f"XLSX-Import Zeile {idx}",
                    notes="XLSX Import", date=date_val)
                db.add(entry)
                inv = Invoice(user_id=user["sub"], filename=f"xlsx-import-{idx}", vendor=vendor,
                    total_amount=amount, vat_amount=mwst, vat_rate=mwst_satz,
                    date=date_val.strftime("%Y-%m-%d") if date_val else "", raw_text=f"XLSX Import: {beschreibung}",
                    invoice_type=entry_type, invoice_number=inv_nr, payment_method=payment,
                    category=category, processed=True)
                db.add(inv)
                imported += 1
            except Exception as e:
                errors.append(f"Zeile {idx}: {str(e)[:80]}")
        db.commit()
        return {"success": True, "imported": imported, "errors": errors}
    except Exception:
        db.rollback()
        logger.exception("XLSX import failed")
        err(500, "XLSX Import failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: DATEV IMPORT
# ============================================================

@app.post("/bookkeeping/import-datev")
async def import_datev(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Import DATEV Buchungsstapel (.csv, semicolon-separated)."""
    import csv
    content = await file.read()
    text = content.decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text), delimiter=";")
    db = SessionLocal()
    imported = 0
    errors = []
    try:
        for idx, row in enumerate(reader, 1):
            try:
                # DATEV format: Umsatz;Soll/Haben;Konto;Gegenkonto;BU;Belegdatum;Buchungstext;USt
                umsatz_raw = row.get("Umsatz") or row.get("umsatz") or "0"
                sh = row.get("Soll/Haben") or row.get("soll/haben") or "S"
                buchungstext = row.get("Buchungstext") or row.get("buchungstext") or ""
                belegdatum = row.get("Belegdatum") or row.get("belegdatum") or ""
                ust = row.get("USt") or row.get("ust") or "19"

                amount = float(umsatz_raw.replace(",", ".").replace(" ", "").strip() or "0")
                if amount <= 0:
                    continue

                entry_type = "expense" if sh.upper() == "S" else "income"

                # Parse DATEV date: DDMM or DDMMYYYY
                date_val = None
                bd = belegdatum.strip()
                if len(bd) == 4:
                    date_val = parse_date_str_to_datetime(f"20{datetime.now().year % 100}-{bd[2:4]}-{bd[0:2]}")
                elif len(bd) >= 6:
                    date_val = parse_date_str_to_datetime(f"{bd[4:]}-{bd[2:4]}-{bd[0:2]}")
                if not date_val:
                    date_val = datetime.now()

                vat_rate = f"{ust}%"
                rate_f = float(ust.replace(",", ".").strip() or "19")
                vat_amount = round(amount * rate_f / (100 + rate_f), 2)

                entry = CashEntry(user_id=user["sub"], description=buchungstext, vendor=buchungstext[:50] or "DATEV Import",
                    gross_amount=amount, vat_amount=vat_amount, vat_rate=vat_rate, entry_type=entry_type,
                    category="other", payment_method="", reference=f"DATEV-Import Zeile {idx}",
                    notes="DATEV Import", date=date_val)
                db.add(entry)
                inv = Invoice(user_id=user["sub"], filename=f"datev-import-{idx}", vendor=buchungstext[:50] or "DATEV Import",
                    total_amount=amount, vat_amount=vat_amount, vat_rate=vat_rate,
                    date=date_val.strftime("%Y-%m-%d") if date_val else "", raw_text=f"DATEV Import: {buchungstext}",
                    invoice_type=entry_type, invoice_number="", payment_method="",
                    category="other", processed=True)
                db.add(inv)
                imported += 1
            except Exception as e:
                errors.append(f"Zeile {idx}: {str(e)[:80]}")
        db.commit()
        return {"success": True, "imported": imported, "errors": errors}
    except Exception:
        db.rollback()
        logger.exception("DATEV import failed")
        err(500, "DATEV Import failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: PHOTO IMPORT (Handwritten Kassenbuch)
# ============================================================

@app.post("/bookkeeping/import-photo")
async def import_kassenbuch_photo(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Import handwritten Kassenbuch photo — OCR handwriting + table parsing"""
    from autotax.ocr import extract_handwriting_text
    import re as _re

    import gc
    content = await file.read()
    try:
        text = await extract_handwriting_text(content, file.filename or "kassenbuch.jpg")
    finally:
        del content
        gc.collect()
    if not text:
        err(400, "Konnte das Bild nicht lesen. Bitte bessere Qualität verwenden.")

    lines = text.strip().split("\n")
    db = SessionLocal()
    imported = 0
    try:
        for line in lines:
            line = line.strip()
            if not line or len(line) < 4:
                continue
            # Pattern 1: date + description + amount (strict)
            m = _re.search(r"(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s+(.+?)\s+(\d+[.,]\d{2})\s*$", line)
            # Pattern 2: date + separator + description + amount (pipes/slashes from OCR)
            if not m:
                m = _re.search(r"(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s*[|/]?\s*(.+?)\s+[/|]?\s*(\d+[.,]\d{2})", line)
            # Pattern 3: date with spaces (OCR misread: "01 03 26" instead of "01.03.26")
            if not m:
                m = _re.search(r"(\d{1,2}\s\d{1,2}\s\d{2,4})\s+(.+?)\s+(\d+[.,]\d{2})", line)
            # Pattern 4: date + description + amount without decimals (e.g. "50" instead of "50,00")
            if not m:
                m = _re.search(r"(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s+(.+?)\s+(\d{1,6})\s*$", line)
            # Pattern 5: date with dashes (01-03-26)
            if not m:
                m = _re.search(r"(\d{1,2}-\d{1,2}-\d{2,4})\s+(.+?)\s+(\d+[.,]\d{2})", line)
            if not m:
                continue
            datum_raw = m.group(1)
            beschreibung = m.group(2).strip()
            amt_str = m.group(3).replace(",", ".")
            betrag = float(amt_str) if "." in amt_str else float(amt_str)
            if betrag <= 0 or len(beschreibung) < 2:
                continue
            # Normalize separators: space, slash, dash → dot
            parts = datum_raw.replace("/", ".").replace("-", ".").replace(" ", ".").split(".")
            date_str = ""
            if len(parts) == 3:
                d, mo, y = parts
                if len(y) == 2:
                    y = "20" + y
                date_str = f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
            vat_amount = round(betrag * 19 / 119, 2)
            entry = CashEntry(
                user_id=user["sub"],
                description=beschreibung,
                vendor=beschreibung[:50],
                gross_amount=betrag,
                vat_amount=vat_amount,
                vat_rate="19%",
                entry_type="expense",
                category="other",
                payment_method="",
                reference="",
                notes="Kassenbuch Foto Import",
                date=parse_date_str_to_datetime(date_str),
            )
            db.add(entry)
            inv = Invoice(
                user_id=user["sub"],
                filename="kassenbuch-foto",
                vendor=beschreibung[:50],
                total_amount=betrag,
                vat_amount=vat_amount,
                vat_rate="19%",
                date=date_str,
                raw_text=f"Kassenbuch Foto: {beschreibung}",
                invoice_type="expense",
                invoice_number="",
                payment_method="",
                category="other",
                processed=True,
            )
            db.add(inv)
            imported += 1
        db.commit()
        return {"success": True, "imported": imported, "ocr_text": text[:500]}
    except Exception:
        db.rollback()
        logger.exception("Kassenbuch photo import failed")
        err(500, "Import failed")
    finally:
        db.close()


@app.post("/api/import-image")
async def import_image_table(file: UploadFile = File(...), save: bool = False, user: dict = Depends(get_current_user)):
    """Import Kassenbuch table image → OCR → structured rows + CSV.
    Columns: Nr, Datum, Beschreibung, Einnahmen, Ausgaben, Saldo
    Returns JSON rows + CSV string. If save=true, also saves to DB.
    """
    from autotax.ocr import extract_handwriting_text, extract_image_text, extract_pdf_text, extract_pdf_page_as_image, extract_table_text
    import re as _re
    import gc

    content = await file.read()
    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()

    # PDF support
    text = ""
    try:
        if "pdf" in content_type or filename.endswith(".pdf"):
            text = extract_pdf_text(content)
            if not text or len(text.strip()) < 20:
                img_bytes = extract_pdf_page_as_image(content)
                if img_bytes:
                    text = await extract_image_text(img_bytes, "scanned.png")
        else:
            # Image: try table-specific OCR first, fallback to printed
            text = await extract_table_text(content, file.filename or "kassenbuch.jpg")

            # If handwriting OCR returned little, try printed OCR (skip if already good enough)
            if not text or len(text.strip()) < 30:
                text_printed = await extract_image_text(content, file.filename or "kassenbuch.png")
                if text_printed and len(text_printed.strip()) > len((text or "").strip()):
                    text = text_printed
    finally:
        del content
        gc.collect()

    if not text or len(text.strip()) < 10:
        err(400, "Konnte das Bild nicht lesen. Bitte bessere Qualität verwenden.")

    import time as _time
    _t0 = _time.time()

    # Preserve raw table before any modification
    raw_lines = [l.strip() for l in text.strip().split("\n") if l.strip()]

    # Amount pattern: matches 42,50 | 1.234,56 | 800,00 | 42.50 | 800 | -16,60
    _AMT_PAT = r"-?\d[\d.]*[.,]\d{1,2}"
    # Loose amount: also matches whole numbers (50, 800) common in handwriting
    _AMT_PAT_LOOSE = r"-?\d[\d.]*(?:[.,]\d{1,2})?"

    # Pre-split: force newline before every date pattern
    _DATE_PAT = (
        r"\d{4}-\d{2}-\d{2}"            # 2026-03-05 (ISO)
        r"|\d{1,2}[./]\d{1,2}[./]\d{2,4}"  # 05.03.2026 or 05/03/26
        r"|\d{1,2}-\d{1,2}-\d{2,4}"      # 05-03-2026
        r"|\d{1,2}[.]\d{1,2}\s\d{2,4}"   # 31.8 21 (OCR broken dot+space)
        r"|\d{1,2}\s\d{1,2}\s\d{2,4}"    # 05 03 2026 (OCR space)
    )
    text = _re.sub(r"(" + _DATE_PAT + r")", r"\n\1", text)

    # Count dates to decide strategy
    all_dates_in_text = _re.findall(_DATE_PAT, text)
    expected_count = len(all_dates_in_text)
    logger.info("Date detection: found %d dates in text (first 3: %s)", expected_count, all_dates_in_text[:3])

    lines = [l.strip() for l in text.strip().split("\n") if l.strip() and len(l.strip()) > 4]
    if len(lines) > 200:
        lines = lines[:200]
    is_table_mode = expected_count > 1
    logger.info("Table import: %d lines, %d dates, table_mode=%s", len(lines), expected_count, is_table_mode)
    rows = []

    def _parse_date(raw):
        raw = raw.strip()
        # Already ISO: 2026-03-05
        if _re.match(r"^\d{4}-\d{2}-\d{2}$", raw):
            return raw
        # Normalize separators: space, dash, slash → dot
        normalized = raw.replace("/", ".").replace("-", ".").replace(" ", ".")
        parts = normalized.split(".")
        if len(parts) == 3:
            dd, mm, yy = parts[0].strip(), parts[1].strip(), parts[2].strip()
            if len(yy) == 2:
                yy = "20" + yy
            if len(dd) <= 2 and len(mm) <= 2 and len(yy) == 4:
                return f"{yy}-{mm.zfill(2)}-{dd.zfill(2)}"
        return raw

    def _detect_currency(t):
        tu = t.upper()
        if "₺" in t or " TL" in tu or tu.endswith("TL"):
            return "TRY"
        if "€" in t or "EUR" in tu:
            return "EUR"
        if "$" in t or "USD" in tu:
            return "USD"
        if "£" in t or "GBP" in tu:
            return "GBP"
        if "CHF" in tu:
            return "CHF"
        return "EUR"  # default for German tax tool

    detected_currency = _detect_currency(text)
    logger.info("Table import currency: %s", detected_currency)

    def _is_date_fragment(s):
        """Check if string looks like a date fragment (DD.MM or MM.YY), not an amount."""
        s = s.strip()
        # DD.MM pattern: 01.01 - 31.12
        m = _re.match(r"^(\d{1,2})[.](\d{1,2})$", s)
        if m:
            d, mo = int(m.group(1)), int(m.group(2))
            if 1 <= d <= 31 and 1 <= mo <= 12:
                return True
        return False

    def _parse_amount(s):
        """Parse German/Turkish number format: 1.234,56 → 1234.56, -16,60 → 16.60"""
        try:
            if "%" in s:
                return 0.0
            raw = s
            s = s.replace("€", "").replace("₺", "").replace(" ", "").strip()
            negative = s.startswith("-")
            s = s.lstrip("-")
            if s.upper().endswith("TL"):
                s = s[:-2].strip()
            if s.upper().endswith("EUR"):
                s = s[:-3].strip()
            if not s:
                return 0.0
            # Reject date-like values: 22.06, 01.12 etc
            if _is_date_fragment(s):
                return 0.0
            # German format: 1.234,56 — dot is thousands, comma is decimal
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                s = s.replace(",", ".")
            val = abs(float(s))  # always positive — sign indicates direction not value
            # Skip year-like numbers
            if _re.match(r"^(19|20)\d{2}$", str(int(val))) and "," not in raw and "." not in raw:
                return 0.0
            return val
        except (ValueError, AttributeError):
            return 0.0

    def _score_line(line):
        """Score a line for table-row likelihood: +2 date, +2 amount, +1 text."""
        score = 0
        if _re.search(_DATE_PAT, line):
            score += 2
        amounts = [a for a in _re.findall(r"(" + _AMT_PAT + r")", line) if not _is_date_fragment(a)]
        if amounts:
            score += 2
        elif _re.search(r"\b\d{2,5}\b", line):
            score += 1  # whole number — weaker signal
        text_part = _re.sub(_DATE_PAT, "", line)
        text_part = _re.sub(_AMT_PAT_LOOSE, "", text_part).strip()
        if len(text_part) >= 3:
            score += 1
        return score

    # Detect column order from header line (e.g. "Nr Datum Beschreibung Einnahmen Ausgaben Saldo")
    _col_order = []  # list of column names in order, e.g. ["einnahmen", "ausgaben", "saldo"]
    for _hl in lines[:5]:
        _hl_lower = _hl.lower()
        if "einnahmen" in _hl_lower or "ausgaben" in _hl_lower:
            # Extract column names in order of appearance
            _header_cols = []
            for _word in _re.findall(r"[a-zäöü]+", _hl_lower):
                if _word in ("einnahmen", "ausgaben", "saldo"):
                    _header_cols.append(_word)
            if _header_cols:
                _col_order = _header_cols
                logger.info("Detected column order from header: %s", _col_order)
            break

    def _assign_amounts_by_columns(amounts_list):
        """Assign amounts to einnahmen/ausgaben based on detected column order."""
        einnahmen, ausgaben = 0.0, 0.0
        if not _col_order:
            # No header detected — use default: 1 amount=ausgaben, 2 amounts=einnahmen+ausgaben
            if len(amounts_list) >= 2:
                einnahmen = amounts_list[0]
                ausgaben = amounts_list[1]
            elif len(amounts_list) == 1:
                ausgaben = amounts_list[0]
            return einnahmen, ausgaben
        # Map amounts to columns by position (skip saldo)
        col_idx = 0
        for col_name in _col_order:
            if col_name == "saldo":
                continue  # always skip saldo
            if col_idx < len(amounts_list):
                if col_name == "einnahmen":
                    einnahmen = amounts_list[col_idx]
                elif col_name == "ausgaben":
                    ausgaben = amounts_list[col_idx]
                col_idx += 1
        return einnahmen, ausgaben

    # Strategy 1: Window-based — group date line + next 2 lines into one entry
    i = 0
    while i < len(lines):
        if len(rows) >= expected_count and expected_count > 0:
            break
        line = lines[i].strip()
        if not line or len(line) < 4:
            i += 1
            continue
        line_lower = line.lower()
        has_date = bool(_re.search(_DATE_PAT, line))
        if not has_date and any(w in line_lower for w in ["datum", "beschreibung", "einnahmen", "ausgaben", "kassenbuch", "übertrag", "seitensumme"]):
            i += 1
            continue

        # If line has date, combine with next 1-2 lines for context
        if has_date:
            combined = line
            lines_consumed = 1
            for j in range(1, 3):
                if i + j < len(lines):
                    next_line = lines[i + j].strip()
                    # Stop if next line starts with a new date
                    if _re.search(r"^(" + _DATE_PAT + r")", next_line):
                        break
                    combined += " " + next_line
                    lines_consumed += 1
            line = combined
            i += lines_consumed
        else:
            i += 1

        # Extract all amounts from line, filter out Saldo (negative/very large running totals)
        _line_amounts_raw = _re.findall(r"(-?\d[\d.]*[.,]\d{1,2})", line)
        _line_amounts = []
        for _a in _line_amounts_raw:
            if _is_date_fragment(_a.lstrip("-")):
                continue
            _v = _parse_amount(_a)
            # Skip negative values (Saldo column) and very large values (>50000 = likely Saldo)
            if _a.strip().startswith("-"):
                continue
            if _v > 50000:
                continue
            _line_amounts.append(_v)

        # Pattern: Date + Description + amounts (Einnahmen/Ausgaben, ignoring Saldo)
        m = _re.search(
            r"(?:\d{1,3}[.\s])?\s*(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s+(.+?)\s+([\d.,]+)\s*",
            line
        )
        if m and _line_amounts:
            datum_raw, beschreibung = m.group(1), m.group(2).strip()
            # Remove all numbers from description
            beschreibung = _re.sub(r"-?\d[\d.]*[.,]\d{1,2}", "", beschreibung).strip()
            beschreibung = _re.sub(r"\s+", " ", beschreibung).strip(" .,;:-")
            if not beschreibung or len(beschreibung) < 2:
                beschreibung = "Eintrag"
            # Assign amounts based on detected column order (or default)
            einnahmen, ausgaben = _assign_amounts_by_columns(_line_amounts)
            if beschreibung and len(beschreibung) >= 2:
                rows.append({"date": _parse_date(datum_raw), "description": beschreibung[:80], "income": round(einnahmen, 2), "expense": round(ausgaben, 2)})
            continue

        # Pattern: DD.MM.YYYY + Description + single amount (no other amounts on line)
        m2 = _re.search(r"(?:\d{1,3}[.\s])?\s*(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s+(.+?)\s+([\d.,]+)\s*$", line)
        if m2:
            datum_raw, beschreibung = m2.group(1), m2.group(2).strip()
            _val = _parse_amount(m2.group(3))
            if beschreibung and len(beschreibung) >= 2 and _val > 0 and _val <= 50000:
                rows.append({"date": _parse_date(datum_raw), "description": beschreibung, "income": 0, "expense": round(_val, 2)})
            continue

        # Pattern: YYYY-MM-DD + Description + two amounts
        m3 = _re.search(r"(\d{4}-\d{2}-\d{2})\s+(.+?)\s+([\d.,]+)\s+([\d.,]+)\s*$", line)
        if m3:
            date_iso, beschreibung = m3.group(1), m3.group(2).strip()
            val1, val2 = _parse_amount(m3.group(3)), _parse_amount(m3.group(4))
            einnahmen = val1 if val1 > 0 and val2 > 0 else 0
            ausgaben = val2 if val1 == 0 or (val1 > 0 and val2 > 0) else val1
            if einnahmen == 0 and ausgaben == 0:
                ausgaben = max(val1, val2)
            if beschreibung and len(beschreibung) >= 2:
                rows.append({"date": date_iso, "description": beschreibung, "income": round(einnahmen, 2), "expense": round(ausgaben, 2)})
            continue

        # Pattern: YYYY-MM-DD + Description + single amount
        m4 = _re.search(r"(\d{4}-\d{2}-\d{2})\s+(.+?)\s+([\d.,]+)\s*$", line)
        if m4:
            date_iso, beschreibung = m4.group(1), m4.group(2).strip()
            if beschreibung and len(beschreibung) >= 2:
                rows.append({"date": date_iso, "description": beschreibung, "income": 0, "expense": round(_parse_amount(m4.group(3)), 2)})
            continue

        # Universal fallback: any date + any text + any amount anywhere in line
        if is_table_mode and has_date:
            date_m = _re.search(r"(" + _DATE_PAT + r")", line)
            if date_m:
                d = date_m.group(1)
                desc = _re.sub(_DATE_PAT, "", line)
                raw_nums = _re.findall(r"(-?" + _AMT_PAT + r")", desc)
                # Filter: no date fragments, no negatives (Saldo), no >50000
                numbers = []
                for n in raw_nums:
                    if _is_date_fragment(n.lstrip("-")):
                        continue
                    if n.strip().startswith("-"):
                        continue
                    pv = _parse_amount(n)
                    if 0 < pv <= 50000:
                        numbers.append(n)
                desc = _re.sub(r"-?" + _AMT_PAT, "", desc).strip()
                desc = _re.sub(r"\s+", " ", desc).strip(" .,;:-")
                if len(desc) < 2:
                    desc = "Eintrag"
                _parsed_nums = [_parse_amount(n) for n in numbers]
                einnahmen, ausgaben = _assign_amounts_by_columns(_parsed_nums)
                parsed_d = _parse_date(d) if "." in d or "/" in d else d
                rows.append({"date": parsed_d, "description": desc[:80], "income": round(einnahmen, 2), "expense": round(ausgaben, 2), "is_uncertain": ausgaben == 0 and einnahmen == 0})

    logger.info("Strategy 1 result: %d rows from %d lines (dates=%d) in %.2fs", len(rows), len(lines), len(all_dates_in_text), _time.time()-_t0)

    # If multiple dates but Strategy 1 found fewer rows → discard and retry
    s1_rows = list(rows)

    # If Strategy 1 got less than expected, try additional strategies
    if is_table_mode and len(rows) < expected_count:
        logger.info("Strategy 1 incomplete: %d/%d — running Strategy 2", len(rows), expected_count)

    # Strategy 2: Try merging split lines (OCR puts dates and amounts on separate lines)
    if is_table_mode and len(rows) < expected_count:
        date_lines = []
        amount_lines = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if _re.match(r"^\d{1,2}[./]\d{1,2}[./]\d{2,4}\s+\S", line) or _re.match(r"^\d{4}-\d{2}-\d{2}\s+\S", line):
                date_lines.append(line)
            elif _re.match(r"^[\d.,]+\s+[\d.,]+\s*$", line):
                amount_lines.append(line)
            elif _re.match(r"^[\d.,]+\s*$", line):
                amount_lines.append(line + " 0")

        logger.info("Table split-line parse: %d date_lines, %d amount_lines", len(date_lines), len(amount_lines))

        for i, dline in enumerate(date_lines):
            m = _re.match(r"(?:\d{1,3}[.\s])?\s*(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s+(.+)", dline)
            if not m:
                continue
            datum_raw, beschreibung = m.group(1), m.group(2).strip()
            beschreibung = _re.sub(r"[\d.,]+\s*$", "", beschreibung).strip()
            if not beschreibung or len(beschreibung) < 2:
                continue
            einnahmen, ausgaben = 0.0, 0.0
            if i < len(amount_lines):
                nums = _re.findall(r"[\d.,]+", amount_lines[i])
                if len(nums) >= 2:
                    v1, v2 = _parse_amount(nums[0]), _parse_amount(nums[1])
                    if v1 > 0 and v2 == 0:
                        ausgaben = v1
                    elif v1 == 0 and v2 > 0:
                        einnahmen = v2
                    elif v1 > 0 and v2 > 0:
                        ausgaben, einnahmen = v1, v2
                elif len(nums) == 1:
                    ausgaben = _parse_amount(nums[0])
            rows.append({"date": _parse_date(datum_raw), "description": beschreibung, "income": round(einnahmen, 2), "expense": round(ausgaben, 2)})

    # Strategy 2.5: Split text by date boundaries (when OCR merges lines)
    if is_table_mode and len(rows) < expected_count:
        logger.info("Strategy 2.5: date-split with %d dates", len(all_dates_in_text))
        # Split text at each date occurrence
        blocks = _re.split(r"(?=\d{1,2}[./]\d{1,2}[./]\d{2,4})|(?=\d{4}-\d{2}-\d{2})", text)
        blocks = [b.strip() for b in blocks if b.strip() and len(b.strip()) > 5][:100]
        logger.info("Strategy 2.5 blocks: %d (first: %s)", len(blocks), blocks[0][:60] if blocks else "none")
        for block in blocks:
            if len(rows) >= expected_count and expected_count > 0:
                break
            # Normalize: merge multi-line block into single line
            block = " ".join(block.splitlines()).strip()
            if not block or len(block) < 5:
                continue
            block_lower = block.lower()
            if any(w in block_lower for w in ["kassenbuch", "übertrag", "seitensumme"]):
                continue

            # Extract date from start of block
            dm = _re.match(r"(\d{1,2}[./]\d{1,2}[./]\d{2,4}|\d{4}-\d{2}-\d{2})\s*(.*)", block)
            if not dm:
                logger.debug("Strategy 2.5 skip (no date): %s", block[:60])
                continue
            date_raw = dm.group(1)
            rest = dm.group(2).strip()

            # Extract amounts from rest (filter out date fragments)
            amounts = [a for a in _re.findall(r"(" + _AMT_PAT + r")", rest) if not _is_date_fragment(a)]
            desc = _re.sub(_AMT_PAT, "", rest).strip()
            desc = _re.sub(r"\s+", " ", desc).strip(" .,;:-")

            if not desc or len(desc) < 2:
                desc = "Eintrag"

            parsed_date = _parse_date(date_raw) if ("." in date_raw or "/" in date_raw) else date_raw
            # Strict positional: numbers[-2]=expense, numbers[-1]=saldo
            if len(amounts) >= 2:
                expense = _parse_amount(amounts[-2])
            elif len(amounts) == 1:
                expense = _parse_amount(amounts[0])
            else:
                expense = 0
            rows.append({"date": parsed_date, "description": desc[:80], "income": 0, "expense": round(expense, 2), "is_uncertain": expense == 0})

        if rows:
            logger.info("Strategy 2.5 date-split: %d rows extracted", len(rows))

    # Strategy 3: ONLY if 0-1 dates found — treat as single receipt
    if not rows and len(all_dates_in_text) <= 1:
        try:
            from autotax.parser import parse_invoice
            parsed = parse_invoice(text)
            vendor = parsed.get("vendor", "")
            amount = parsed.get("total_amount", 0)
            date_str = parsed.get("date", "")
            if amount is not None:
                rows.append({
                    "date": date_str or datetime.now().strftime("%Y-%m-%d"),
                    "description": vendor or "Beleg",
                    "income": 0,
                    "expense": round(float(amount), 2),
                })
                logger.info("Table import fallback: single receipt %s €%.2f", vendor, amount)
        except Exception:
            pass

    # Strategy 5: Column-based parsing (OCR reads columns separately: dates, descriptions, amounts)
    # Detects when OCR returns all dates first, then all descriptions, then all amounts
    valid_amounts_in_rows = sum(1 for r in rows if r.get("expense", 0) > 0 or r.get("income", 0) > 0)
    if is_table_mode and valid_amounts_in_rows < expected_count // 2:
        try:
            _skip = {"einnahmen", "ausgaben", "beschreibung", "datum", "nr.", "mwst", "brutto", "netto", "summe", "kassenbuch", "saldo", "ubersicht"}
            col_dates, col_descs, col_amounts = [], [], []
            for line in raw_lines:
                ll = line.lower().strip()
                if any(w in ll for w in _skip):
                    continue
                l = line.strip()
                # "21 26.8.21" → row number + date merged
                merged = _re.match(r"^\d{1,3}\s+(" + _DATE_PAT + r")", l)
                if merged:
                    col_dates.append(merged.group(1))
                elif _re.match(r"^(" + _DATE_PAT + r")", l):
                    col_dates.append(l)
                elif _re.match(r"^" + _AMT_PAT + r"\s*[€$₺]?$", l):
                    col_amounts.append(l)
                elif _re.match(r"^-\s+\d", l):
                    col_amounts.append(l.replace(" ", ""))  # "- 29,28" → "-29,28"
                elif _re.match(r"^\d{1,3}$", l):
                    continue
                elif len(line.strip()) > 2 and not _re.match(r"^[\d.,€$₺/ \-]+$", line.strip()):
                    col_descs.append(line.strip())

            # Filter amounts: skip negative (saldo), keep positive (expense)
            expenses_only = [a for a in col_amounts if not a.startswith("-")]

            logger.info("Strategy 5 columns: %d dates, %d descs, %d amounts (%d positive)",
                        len(col_dates), len(col_descs), len(col_amounts), len(expenses_only))

            n = min(len(col_dates), len(col_descs), len(expenses_only))
            if n > len(rows) // 2:  # only use if significantly better
                s5_rows = []
                for i in range(n):
                    # Extract date from date line (might have extra text)
                    dm = _re.search(r"(" + _DATE_PAT + r")", col_dates[i])
                    if not dm:
                        continue
                    date_val = _parse_date(dm.group(1))
                    desc = _re.sub(_DATE_PAT, "", col_dates[i]).strip()
                    if not desc or len(desc) < 2:
                        desc = col_descs[i] if i < len(col_descs) else "Eintrag"
                    else:
                        desc = desc + " " + (col_descs[i] if i < len(col_descs) else "")
                    desc = desc.strip()[:80]
                    amt = _parse_amount(expenses_only[i]) if i < len(expenses_only) else 0
                    s5_rows.append({"date": date_val, "description": desc, "income": 0, "expense": round(amt, 2), "is_uncertain": amt == 0})

                if len(s5_rows) > valid_amounts_in_rows:
                    s5_valid = sum(1 for r in s5_rows if r["expense"] > 0)
                    logger.info("Strategy 5 result: %d rows (%d with amounts) vs current %d (%d with amounts)",
                                len(s5_rows), s5_valid, len(rows), valid_amounts_in_rows)
                    if s5_valid > valid_amounts_in_rows:
                        rows = s5_rows
                        logger.info("Strategy 5 accepted: %d rows", len(rows))
        except Exception as e:
            logger.warning("Strategy 5 column-based failed: %s", e)

    # Strategy 6: Scored fallback — if all table strategies failed, score each line
    # and return partial matches with confidence, rather than empty result
    if not rows and text and len(text.strip()) > 10:
        logger.info("Strategy 6: all table strategies failed, trying scored extraction")
        raw_text_lines = [l.strip() for l in text.strip().split("\n") if l.strip() and len(l.strip()) > 3]
        for rl in raw_text_lines[:100]:
            rl_lower = rl.lower()
            if any(w in rl_lower for w in ["datum", "beschreibung", "einnahmen", "ausgaben", "kassenbuch", "übertrag", "seitensumme", "saldo"]):
                continue
            score = _score_line(rl)
            if score < 2:
                continue  # skip lines with no useful signal
            date_m = _re.search(r"(" + _DATE_PAT + r")", rl)
            date_val = _parse_date(date_m.group(1)) if date_m else ""
            # Try strict amount first, then loose (whole numbers)
            amounts = [a for a in _re.findall(r"(" + _AMT_PAT + r")", rl) if not _is_date_fragment(a)]
            if not amounts:
                amounts = _re.findall(r"\b(\d{2,5})\b", rl)
            desc = _re.sub(_DATE_PAT, "", rl)
            desc = _re.sub(_AMT_PAT_LOOSE, "", desc)
            desc = _re.sub(r"\s+", " ", desc).strip(" .,;:-|/")
            if not desc or len(desc) < 2:
                desc = rl[:80]
            expense = _parse_amount(amounts[0]) if amounts else 0
            confidence = round(min(score / 5.0, 1.0), 2)
            rows.append({
                "date": date_val,
                "description": desc[:80],
                "income": 0,
                "expense": round(expense, 2),
                "is_uncertain": score < 4,
                "confidence": confidence,
                "raw_fallback": True,
            })
        if rows:
            logger.info("Strategy 6 scored fallback: %d lines (avg confidence=%.2f)",
                        len(rows), sum(r.get("confidence", 0) for r in rows) / len(rows))

    # Strategy 7: LLM fallback — if no rows or low confidence, try Claude Haiku
    _avg_conf = (sum(r.get("confidence", 1.0) for r in rows) / len(rows)) if rows else 0
    _high_conf_count = sum(1 for r in rows if r.get("confidence", 1.0) >= 0.8)
    _text_len = len(text.strip()) if text else 0
    _raw_count = len(raw_lines)
    _llm_skip_reason = None
    if _text_len < 20:
        _llm_skip_reason = "hard block: text too short (%d chars)" % _text_len
    elif _text_len <= 50:
        _llm_skip_reason = "low text length (%d chars)" % _text_len
    elif rows and _avg_conf >= 0.7:
        _llm_skip_reason = "high confidence (%.2f)" % _avg_conf
    elif _high_conf_count > 0:
        _llm_skip_reason = "has %d high-confidence rows" % _high_conf_count
    elif _raw_count <= 2:
        _llm_skip_reason = "insufficient raw rows (%d)" % _raw_count
    elif rows and _avg_conf >= 0.6:
        _llm_skip_reason = "adequate confidence (%.2f)" % _avg_conf

    if _llm_skip_reason:
        logger.info("LLM skipped: %s", _llm_skip_reason)
    else:
        try:
            from autotax.ocr import llm_parse_table
            _user_id = str(user.get("sub", ""))
            _user_plan = str(user.get("plan", "free"))
            llm_rows = await llm_parse_table(text, user_id=_user_id, user_plan=_user_plan)
            # Validate LLM rows: must have amount + (date or text)
            llm_rows = [r for r in llm_rows if (r.get("expense", 0) + r.get("income", 0)) > 0 and (r.get("date", "").strip() or len(r.get("description", "").strip()) >= 3)]
            if llm_rows:
                if not rows:
                    rows = llm_rows
                    logger.info("Strategy 7: LLM provided %d rows (no prior rows)", len(llm_rows))
                elif _avg_conf < 0.6 and len(llm_rows) > len(rows):
                    rows = llm_rows
                    logger.info("Strategy 7: LLM provided %d rows (replaced %d low-confidence)", len(llm_rows), len(rows))
        except Exception as e:
            logger.warning("Strategy 7 LLM fallback failed: %s", e)

    # Ensure all rows have is_uncertain flag
    for r in rows:
        if "is_uncertain" not in r:
            r["is_uncertain"] = False

    # Deduplicate: same date + same description = duplicate
    seen = set()
    unique_rows = []
    for r in rows:
        key = (r["date"], r["description"][:30])
        if key not in seen:
            seen.add(key)
            unique_rows.append(r)
    if len(unique_rows) < len(rows):
        logger.info("Dedup: %d → %d rows", len(rows), len(unique_rows))
    rows = unique_rows

    # Data quality: validate rows — reject rows without amount or without (date or text)
    validated = []
    for r in rows:
        amount = r.get("expense", 0) + r.get("income", 0)
        has_amount = amount > 0
        has_date = bool(r.get("date", "").strip())
        has_text = len(r.get("description", "").strip()) >= 3
        if not has_amount:
            continue  # RULE 2: reject if no valid amount
        if not (has_date or has_text):
            continue  # RULE 3: need date or text
        # RULE 1: mark low-confidence rows
        if r.get("confidence", 1.0) < 0.6:
            r["is_uncertain"] = True
        validated.append(r)
    if len(validated) < len(rows):
        logger.info("Validation: %d → %d rows (rejected %d)", len(rows), len(validated), len(rows) - len(validated))
    rows = validated

    logger.info("Table import result: %d rows (expected %d)", len(rows), expected_count)

    # Generate CSV
    csv_lines = ["Datum,Beschreibung,Einnahmen,Ausgaben"]
    for r in rows:
        desc = r["description"].replace('"', '""')
        inc = f'{r["income"]:.2f}' if not r.get("is_uncertain") or r["income"] > 0 else ""
        exp = f'{r["expense"]:.2f}' if not r.get("is_uncertain") or r["expense"] > 0 else ""
        csv_lines.append(f'{r["date"]},"{desc}",{inc},{exp}')
    csv_text = "\n".join(csv_lines)

    # Optionally save to DB — block if data quality too low
    saved = 0
    _save_avg_conf = (sum(r.get("confidence", 1.0) for r in rows) / len(rows)) if rows else 0
    _save_blocked = len(rows) < 2 or _save_avg_conf < 0.5
    if _save_blocked and save:
        logger.info("Auto-save blocked: %d rows, avg confidence=%.2f (min 2 rows, 0.5 confidence)", len(rows), _save_avg_conf)
    if save and rows and not _save_blocked:
        db = SessionLocal()
        try:
            for r in rows:
                amount = r["income"] if r["income"] > 0 else r["expense"]
                entry_type = "income" if r["income"] > 0 else "expense"
                vat_amount = round(amount * 19 / 119, 2) if amount > 0 else 0
                date_val = parse_date_str_to_datetime(r["date"])
                if not date_val:
                    date_val = datetime.now()
                entry = CashEntry(
                    user_id=user["sub"],
                    description=r["description"],
                    vendor=r["description"][:50],
                    gross_amount=amount,
                    vat_amount=vat_amount,
                    vat_rate="19%",
                    entry_type=entry_type,
                    category="other",
                    payment_method="",
                    reference="",
                    notes="Bild Import",
                    date=date_val,
                )
                db.add(entry)
                inv = Invoice(
                    user_id=user["sub"],
                    filename="bild-import",
                    vendor=r["description"][:50],
                    total_amount=amount,
                    vat_amount=vat_amount,
                    vat_rate="19%",
                    date=r["date"],
                    raw_text=f"Bild Import: {r['description']}",
                    invoice_type=entry_type,
                    invoice_number="",
                    payment_method="",
                    category="other",
                    processed=True,
                )
                db.add(inv)
                saved += 1
            db.commit()
        except Exception:
            db.rollback()
            logger.exception("Image table import save failed")
        finally:
            db.close()

    # Add currency to all rows
    for r in rows:
        r["currency"] = detected_currency

    logger.info("Table import complete: %d rows in %.2fs", len(rows), _time.time()-_t0)

    return {
        "success": True,
        "rows": rows,
        "row_count": len(rows),
        "saved": saved,
        "save_blocked": _save_blocked if save else False,
        "csv": csv_text,
        "currency": detected_currency,
        "raw_rows": raw_lines,
        "raw_row_count": len(raw_lines),
        "ocr_text": text[:2000],
    }



# ============================================================
# TAX: EÜR
# ============================================================

@app.get("/tax/euer")
def list_euer(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"]).all()
        years = set()
        for i in invoices:
            d = safe_date_str(i.date)
            if len(d) >= 4:
                years.add(d[:4])
        result = []
        for y in sorted(years):
            year_invs = [i for i in invoices if safe_date_str(i.date).startswith(y)]
            einnahmen = sum(safe_float(i.total_amount) for i in year_invs if safe_invoice_type(i.invoice_type) == "income")
            ausgaben = sum(safe_float(i.total_amount) for i in year_invs if safe_invoice_type(i.invoice_type) == "expense")
            result.append({
                "id": int(y),
                "steuerjahr": int(y),
                "summe_einnahmen": round(einnahmen, 2),
                "summe_ausgaben": round(ausgaben, 2),
                "gewinn_verlust": round(einnahmen - ausgaben, 2),
            })
        return result
    except Exception:
        logger.exception("EÜR list failed")
        err(500, "Failed")
    finally:
        db.close()


@app.post("/tax/euer/auto-fill")
def auto_fill_euer(steuerjahr: int = Query(...), user: dict = Depends(get_current_user)):
    return {"success": True, "steuerjahr": steuerjahr, "status": "generated"}


# ============================================================
# CHAT
# ============================================================


@app.post("/feedback")
def submit_feedback(body: dict = Body(...), user: dict = Depends(get_current_user)):
    message = body.get("message", "")
    if not message.strip():
        err(400, "Feedback message is empty")
    logger.info("FEEDBACK from user %s: %s", user.get("email", user["sub"]), message[:500])
    return {"success": True, "message": "Feedback received"}


# ============================================================
# COMPANIES (max 2 per user)
# ============================================================
# RECEIPT VAULT
# ============================================================

@app.get("/vault")
def list_vault(search: Optional[str] = Query(None), user: dict = Depends(get_current_user)):
    """List all receipts with metadata — checks DB for original file."""
    db = SessionLocal()
    try:
        q = db.query(Invoice).filter(Invoice.user_id == user["sub"])
        q = q.filter((Invoice.is_deleted == False) | (Invoice.is_deleted == None))
        if search:
            from sqlalchemy import or_
            q = q.filter(or_(Invoice.vendor.ilike(f"%{search}%"), Invoice.date.ilike(f"%{search}%")))
        invoices = q.order_by(Invoice.created_at.desc()).all()
        items = []
        for inv in invoices:
            items.append({
                "id": inv.id,
                "vendor": safe_vendor(inv.vendor),
                "date": safe_date_str(inv.date),
                "total_amount": safe_float(inv.total_amount),
                "vat_amount": safe_float(inv.vat_amount),
                "category": safe_category(inv.category),
                "filename": inv.filename or "",
                "has_original": inv.file_data is not None and len(inv.file_data) > 0 if inv.file_data else False,
                "file_content_type": inv.file_content_type or "",
                "invoice_type": safe_invoice_type(inv.invoice_type),
            })
        return {"items": items, "total": len(items)}
    finally:
        db.close()


# --- ADDED START: Upload/replace original file for existing invoice ---
@app.post("/vault/{invoice_id}/upload")
async def upload_vault_file(invoice_id: int, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload or replace original receipt file for an existing invoice."""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            err(400, "Datei zu groß")
        inv.file_data = content
        inv.file_content_type = file.content_type or "application/octet-stream"
        inv.filename = file.filename or inv.filename
        db.commit()
        logger.info("Vault upload: invoice %d, %d bytes", invoice_id, len(content))
        return {"success": True}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Vault upload failed")
        err(500, "Upload failed")
    finally:
        db.close()
# --- ADDED END ---


@app.get("/vault/{invoice_id}/download")
def download_vault_file(invoice_id: int, mode: str = Query("inline"), user: dict = Depends(get_current_user)):
    """Download original receipt file from DB. mode=inline (preview) or attachment (download)."""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Not found")
        if not inv.file_data:
            err(404, "Kein Original gespeichert")
        ct = inv.file_content_type or "application/octet-stream"
        fname = inv.filename or "beleg"
        disposition = "attachment" if mode == "attachment" else "inline"
        return StreamingResponse(io.BytesIO(inv.file_data), media_type=ct, headers={"Content-Disposition": f"{disposition}; filename={fname}"})
    finally:
        db.close()


# ============================================================
# PRICING & PLAN
# ============================================================

PRICING = {
    "free": {"name": "Free", "price": 0, "max_invoices": 50, "max_companies": 2},
    "early": {"name": "Early Adopter", "price": 10, "max_invoices": 500, "max_companies": 5},
    "pro": {"name": "Pro", "price": 20, "max_invoices": -1, "max_companies": -1},
}


@app.get("/pricing")
def get_pricing():
    return {"plans": PRICING}


@app.get("/account/plan")
def get_user_plan(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user["sub"]).first()
        if not u:
            err(404, "User not found")
        plan = u.plan or "free"
        inv_count = db.query(Invoice).filter(Invoice.user_id == user["sub"]).count()
        plan_info = PRICING.get(plan, PRICING["free"])
        return {
            "plan": plan,
            "plan_name": plan_info["name"],
            "price": plan_info["price"],
            "max_invoices": plan_info["max_invoices"],
            "invoice_count": inv_count,
            "is_early": plan == "early",
            "message": "Frühe Nutzer behalten ihren Preis" if plan == "early" else None,
        }
    finally:
        db.close()


@app.post("/account/upgrade")
def upgrade_plan(body: dict = Body(...), user: dict = Depends(get_current_user)):
    plan = body.get("plan", "pro")
    if plan not in ("early", "pro"):
        err(400, "Invalid plan")
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user["sub"]).first()
        if not u:
            err(404, "User not found")
        u.plan = plan
        db.commit()
        return {"success": True, "plan": plan}
    finally:
        db.close()


# ============================================================

@app.get("/companies")
def list_companies(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        companies = db.query(UserCompany).filter(UserCompany.user_id == user["sub"]).all()
        return [{"id": c.id, "company_name": c.company_name, "iban": c.iban or "", "tax_id": c.tax_id or "", "address": c.address or "", "phone": c.phone or "", "fax": c.fax or "", "email": c.email or "", "website": c.website or ""} for c in companies]
    finally:
        db.close()


@app.post("/companies")
def add_company(body: dict = Body(...), user: dict = Depends(get_current_user)):
    company_name = body.get("company_name", "").strip()
    if not company_name:
        err(400, "Firmenname ist erforderlich")
    db = SessionLocal()
    try:
        existing = db.query(UserCompany).filter(UserCompany.user_id == user["sub"]).count()
        if existing >= 2:
            err(400, "Maximal 2 Firmen erlaubt. Upgrade auf Pro für mehr.")
        dup = db.query(UserCompany).filter(UserCompany.user_id == user["sub"], UserCompany.company_name == company_name).first()
        if dup:
            err(400, "Firma existiert bereits")
        c = UserCompany(
            user_id=user["sub"], company_name=company_name,
            iban=body.get("iban", "").strip() or None,
            tax_id=body.get("tax_id", "").strip() or None,
            address=body.get("address", "").strip() or None,
            phone=body.get("phone", "").strip() or None,
            fax=body.get("fax", "").strip() or None,
            email=body.get("email", "").strip() or None,
            website=body.get("website", "").strip() or None,
        )
        db.add(c)
        db.commit()
        db.refresh(c)
        return {"success": True, "id": c.id, "company_name": c.company_name}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Add company failed")
        err(500, "Failed")
    finally:
        db.close()


@app.delete("/companies/{company_id}")
def delete_company(company_id: int, user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        c = db.query(UserCompany).filter(UserCompany.id == company_id, UserCompany.user_id == user["sub"]).first()
        if not c:
            err(404, "Firma nicht gefunden")
        db.delete(c)
        db.commit()
        return {"success": True}
    finally:
        db.close()


# ============================================================

@app.post("/chat")
def chat_endpoint(body: dict = Body(...), user: dict = Depends(get_current_user)):
    message = body.get("message", "")
    db = SessionLocal()
    try:
        all_invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"]).all()
        # Filter invalid entries — same logic as dashboard
        invoices = [i for i in all_invoices if safe_float(i.total_amount) > 0 and safe_vendor(i.vendor) != "Unbekannt"]

        inv_count = len(invoices)
        inv_sum = sum(safe_float(i.total_amount) for i in invoices)

        inv_inc = [i for i in invoices if safe_invoice_type(i.invoice_type) == "income"]
        inv_exp = [i for i in invoices if safe_invoice_type(i.invoice_type) == "expense"]

        total_income = sum(safe_float(i.total_amount) for i in inv_inc)
        total_expenses = sum(safe_float(i.total_amount) for i in inv_exp)
        net_profit = total_income - total_expenses

        vat_paid = sum(safe_float(i.vat_amount) for i in inv_exp)
        vat_collected = sum(safe_float(i.vat_amount) for i in inv_inc)
        vat_balance = vat_collected - vat_paid

        cat_map = {}
        for i in invoices:
            c = safe_category(i.category)
            cat_map[c] = cat_map.get(c, 0) + safe_float(i.total_amount)
        cat_str = ", ".join(f"{k}: €{v:.2f}" for k, v in sorted(cat_map.items(), key=lambda x: -x[1])) if cat_map else "keine"

        vendors = {}
        for i in invoices:
            v = safe_vendor(i.vendor)
            vendors[v] = vendors.get(v, 0) + safe_float(i.total_amount)
        top_vendors = ", ".join(f"{k}: €{v:.2f}" for k, v in sorted(vendors.items(), key=lambda x: -x[1])[:5]) if vendors else "keine"

        msg = message.lower().strip()
        reply = None

        # ── TIER 1: Multi-word / very specific phrases (must match BEFORE single words) ──

        # Rechnung erstellen (before generic "rechnung")
        if any(w in msg for w in ["rechnung erstellen", "rechnung schreiben", "rechnung anlegen", "neue rechnung", "fatura oluştur", "fatura yaz"]):
            reply = "🧾 Rechnung erstellen:\n\nSo geht's:\n1. Gehe zu Rechnungen → '+ Rechnung erstellen' Button\n2. Felder ausfüllen: Kunde, Betrag, MwSt-Satz, Datum, Rechnungs-Nr.\n3. 'Speichern' klicken → Wird als Einnahme gespeichert\n4. 'PDF herunterladen?' Dialog → Rechnung als PDF speichern\n\n• MwSt wird automatisch berechnet (19%, 7%, oder 0%)\n• Kleinunternehmer: §19-Hinweis erscheint automatisch auf PDF\n• PDF enthält: Firmenname, Kunde, Betrag, MwSt, Netto/Brutto\n• Tipp: Rechnungs-Nr. vergeben (z.B. RE-2026-001)"

        # E-Rechnung (before generic "rechnung")
        elif any(w in msg for w in ["e-rechnung", "erechnung", "xrechnung", "zugferd", "xml rechnung", "elektronische rechnung", "factur-x"]):
            reply = "📄 E-Rechnung (XRechnung / ZUGFeRD):\n\nSo geht's:\n1. Gehe zu Upload\n2. Klicke 'E-Rechnung (XML) importieren'\n3. XML-Datei auswählen → Automatische Erkennung\n\n• XRechnung (XML) und ZUGFeRD (PDF) werden unterstützt\n• Automatische Erkennung: Lieferant, Betrag, MwSt, Datum, Rechnungs-Nr., Kategorie, DATEV-Konto\n• Seit 01.01.2025 Pflicht für B2B in Deutschland (Empfang)\n• Auch Kleinunternehmer müssen E-Rechnungen empfangen können\n• Unterstützte Formate: XRechnung, ZUGFeRD, Factur-X"

        # Kleinunternehmer (before generic "steuer")
        elif any(w in msg for w in ["kleinunternehmer", "§19", "paragraph 19", "keine mwst", "keine umsatzsteuer", "küçük işletme", "klein unternehmer"]):
            reply = "📋 Kleinunternehmerregelung (§19 UStG):\n\n• Profil-Menü (oben rechts) → Kleinunternehmer Toggle aktivieren\n• Keine MwSt auf erstellten Rechnungen\n• Hinweis 'Gemäß §19 UStG wird keine Umsatzsteuer berechnet' erscheint automatisch auf PDF\n• Grenze: 25.000€ Umsatz im Vorjahr / 100.000€ laufendes Jahr\n• E-Rechnungen empfangen ist trotzdem Pflicht (seit 2025)\n• Vorteil: Weniger Bürokratie, keine USt-Voranmeldung"

        # EÜR (before generic "steuer")
        elif any(w in msg for w in ["eür", "einnahmen-überschuss", "überschussrechnung", "steuerformular", "steuererklärung"]):
            reply = "🧾 EÜR (Einnahmen-Überschuss-Rechnung):\n\nSo geht's:\n1. Sidebar → 'Steuer (EÜR)'\n2. Steuerjahr wählen\n3. 'Generieren' klicken\n\n• EÜR = Einnahmen-Überschuss-Rechnung (für Freiberufler und Kleinunternehmer)\n• Automatische Berechnung aus allen Rechnungen und Kassenbuch-Einträgen\n• Enthält: Betriebseinnahmen, Betriebsausgaben, Gewinn/Verlust, MwSt-Zusammenfassung\n• Kann als Grundlage für die Steuererklärung verwendet werden\n\nHinweis: Für die offizielle Steuererklärung immer einen Steuerberater konsultieren."

        # DATEV (before generic "export")
        elif any(w in msg for w in ["datev", "steuerberater export", "buchungskonto", "skr03", "skr04", "skr 03", "skr 04"]):
            reply = "📊 DATEV Export:\n\nSo geht's:\n1. Gehe zu Export\n2. Optional: Steuerjahr wählen\n3. 'DATEV' Button klicken → Datei wird heruntergeladen\n\n• Standard-Format für deutsche Steuerberater\n• Automatische Kontenzuordnung:\n  - 6800 = Wareneinkauf / Lebensmittel\n  - 6640 = Bewirtung / Restaurant\n  - 6670 = Kfz-Kosten / Kraftstoff\n  - 6815 = Bürobedarf / Software\n  - 8400 = Erlöse / Einnahmen\n• Dein Steuerberater kann die Datei direkt in DATEV-Software importieren\n• Tipp: Vor dem Export Kategorien prüfen — sie bestimmen die Kontozuordnung"

        # CSV Import (before generic "csv" / "import")
        elif any(w in msg for w in ["csv import", "foto import", "importieren", "içe aktar", "einlesen"]):
            reply = "📥 Import-Optionen:\n\n1. CSV Import:\n• Kassenbuch → 'CSV Import' Button → CSV-Datei auswählen\n• Trennzeichen: Komma oder Semikolon (automatisch erkannt)\n• Spaltenbezeichnungen: Deutsch oder Englisch\n• Datumsformat: DD.MM.YYYY oder YYYY-MM-DD\n• Einnahmen/Ausgaben werden automatisch erkannt\n• MwSt wird automatisch mit 19% berechnet\n• Jede Zeile erstellt Kassenbuch-Eintrag UND Rechnung\n• Beispiel: Datum,Beschreibung,Lieferant,Ausgaben,Einnahmen,Kategorie\n\n2. Foto Import:\n• Kassenbuch → 'Foto Import' Button\n• OCR erkennt handgeschriebene Tabellen: Datum | Beschreibung | Betrag\n• Tipp: Gerade und bei guter Beleuchtung fotografieren\n\n3. Beleg Upload:\n• Upload-Seite → PDF/Foto hochladen"

        # PDF / Drucken (before generic "download")
        elif any(w in msg for w in ["drucken", "pdf", "ausdrucken", "yazdır", "print"]):
            reply = "🖨️ PDF Drucken:\n\nWo findest du den PDF Button?\n• Rechnungen → 🖨️ PDF Button neben jeder Rechnung\n• Kassenbuch → 🖨️ PDF Button (wenn Rechnung verknüpft)\n• Belege → 🖨️ PDF Button unter jedem Beleg\n• Rechnung erstellen → Nach Speichern: 'PDF herunterladen?' Dialog\n\nPDF enthält:\n• Firmenname und E-Mail (aus deinem Profil)\n• Kunde/Lieferant, Rechnungs-Nr., Datum\n• Beschreibung, Kategorie, MwSt-Satz\n• Netto, MwSt-Betrag, Gesamtbetrag (Brutto)\n• §19-Hinweis (wenn Kleinunternehmer aktiviert)\n• Footer: 'Erstellt mit AutoTax-HUB'"

        # Preise / Pricing (before generic "kosten" catches it)
        elif any(w in msg for w in ["pricing", "abo", "plan ", "upgrade", "tarif", "paket", "ücret"]) or (any(w in msg for w in ["preis", "fiyat", "kosten"]) and any(w in msg for w in ["monat", "plan", "abo", "wie viel kostet", "was kostet autotax"])):
            reply = "💰 Preise:\n• Free: €0/Monat — 50 Rechnungen, 2 Firmen, CSV Export\n• Early Adopter: €10/Monat — 500 Rechnungen, 5 Firmen, DATEV\n• Pro: €20/Monat — Unbegrenzt, API, Priority Support\n• Frühe Nutzer behalten ihren Preis dauerhaft!\n• Stripe-Zahlung kommt bald"

        # Firmen verwalten (before "lieferant" catches "firma")
        elif any(w in msg for w in ["firmen verwalten", "meine firma", "firma ändern", "firma registrier", "firma hinzufügen"]):
            reply = "🏢 Firmen verwalten:\n• Sidebar → 'Firmen' Seite\n• Max. 2 Firmen registrieren\n• Firmenname wird für Einnahme-Erkennung verwendet\n• Upload: Vendor = deine Firma → automatisch Einnahme\n• Firma kann nicht geändert werden (Kontakt Support)"

        # Wie viele / Anzahl (before "wie viel" catches it)
        elif any(w in msg for w in ["wie viele", "anzahl", "count", "kaç tane", "adet"]):
            reply = f"📊 Anzahl:\n• Rechnungen: {inv_count}\n• Einnahmen: {len(inv_inc)}\n• Ausgaben: {len(inv_exp)}"

        # ── TIER 2: Greetings & meta (catch early to avoid false matches) ──

        # Hallo / Greeting
        elif any(w in msg for w in ["hallo", "hey", "merhaba", "hello", "guten tag", "guten morgen", "guten abend", "selam", "servus", "grüß"]) or msg in ["hi", "na"]:
            reply = f"👋 Hallo! Du hast {inv_count} Rechnungen. Wie kann ich dir helfen? Tippe 'Hilfe' für eine Übersicht."

        # Danke
        elif any(w in msg for w in ["danke", "thanks", "thx", "merci", "teşekkür", "sağol", "gracias", "super", "perfekt", "top"]):
            reply = "Gerne! Wenn du weitere Fragen hast, frag einfach. 😊"

        # Hilfe / Help
        elif any(w in msg for w in ["hilfe", "help", "was kannst", "anleitung", "wie funktioniert", "feature", "yardım", "yardim", "nasıl", "nasil", "nedir", "ne yapabilir", "fonksiyon", "what can"]):
            reply = "🤖 Ich kann dir helfen mit:\n• 'Wie viel?' — Gesamtbeträge\n• 'Kategorien' — Ausgaben nach Kategorie\n• 'MwSt' / 'KDV' — Vorsteuer & USt\n• 'Steuer' — Steuerschätzung\n• 'Gewinn' — Einnahmen vs. Ausgaben\n• 'Lieferanten' — Top Anbieter\n• 'Dashboard' — Finanzübersicht\n• 'Kassenbuch' — Kassenbuch-Status\n• 'Rechnungen' — Rechnungsübersicht\n• 'Upload' — Belege hochladen\n• 'Import' — CSV oder Foto importieren\n• 'Export' / 'CSV' / 'DATEV' — Exportieren\n• 'EÜR' — Steuererklärung\n• 'E-Rechnung' — XRechnung / ZUGFeRD\n• 'Rechnung erstellen' — Eigene Rechnungen\n• 'PDF' / 'Drucken' — PDF herunterladen\n• 'Kleinunternehmer' — §19 UStG\n• 'Firmen' — Firmenverwaltung\n• 'Preise' — Pläne & Abo\n• 'App' / 'PWA' — Mobile Nutzung\n• 'Belege' — Belegverwaltung\n• 'Sync' — Synchronisation\n• 'QR' — QR-Code Erkennung\n• 'Passwort' — Konto & Login\n• 'Löschen' — Einträge entfernen\n\nOder frag einfach frei — z.B. einen Lieferanten-Namen!"

        # ── TIER 3: Data queries (user's actual financial data) ──

        # Summe / Gesamt / wie viel
        elif any(w in msg for w in ["wie viel", "wieviel", "summe", "total", "gesamt", "how much", "insgesamt", "ne kadar", "özet", "zusammenfassung", "overview", "toplam"]):
            reply = f"📊 Übersicht:\n• Rechnungen: {inv_count} (€{inv_sum:.2f})\n• Einnahmen: €{total_income:.2f}\n• Ausgaben: €{total_expenses:.2f}\n• Gewinn: €{net_profit:.2f}"

        # Kategorie
        elif any(w in msg for w in ["kategorie", "categories", "aufteilung", "verteilung", "category", "kategori"]):
            reply = f"📂 Kategorien:\n{cat_str}"

        # MwSt / VAT (before generic "steuer")
        elif any(w in msg for w in ["mwst", "vat", "umsatzsteuer", "mehrwertsteuer", "vorsteuer", "kdv", "tva"]):
            reply = f"🧾 MwSt-Übersicht:\n• Gezahlte Vorsteuer: €{vat_paid:.2f}\n• Vereinnahmte USt: €{vat_collected:.2f}\n• Saldo: €{vat_balance:.2f}\n{'→ Du bekommst €'+str(abs(round(vat_balance,2)))+' zurück' if vat_balance < 0 else '→ Du schuldest €'+str(round(vat_balance,2)) if vat_balance > 0 else '→ Ausgeglichen'}"

        # Steuer / Einkommensteuer
        elif any(w in msg for w in ["steuer", "tax", "einkommensteuer", "steuerlast", "gelir vergisi"]):
            if net_profit > 277826:
                rate = 45
            elif net_profit > 61356:
                rate = 42
            elif net_profit > 17005:
                rate = 30
            elif net_profit > 10908:
                rate = 14
            else:
                rate = 0
            estimate = round(net_profit * rate / 100, 2) if net_profit > 0 else 0
            reply = f"💰 Steuer-Schätzung (Deutschland):\n• Gewinn: €{net_profit:.2f}\n• Steuersatz: {rate}%\n• Geschätzte Steuer: €{estimate:.2f}\n\nHinweis: Dies ist eine Schätzung. Für genaue Berechnung bitte Steuerberater konsultieren."

        # Einnahmen / Income
        elif any(w in msg for w in ["einnahme", "income", "umsatz", "revenue", "verdien", "gelir", "kazanç"]):
            reply = f"📈 Einnahmen: €{total_income:.2f} ({len(inv_inc)} Positionen)"

        # Ausgaben / Expenses
        elif any(w in msg for w in ["ausgabe", "expense", "kosten", "cost", "bezahl", "gider", "harcama", "masraf"]):
            reply = f"📉 Ausgaben: €{total_expenses:.2f} ({len(inv_exp)} Positionen)"

        # Gewinn / Profit
        elif any(w in msg for w in ["gewinn", "profit", "verlust", "loss", "ergebnis", "kâr", "kar", "zarar"]):
            emoji = "📈" if net_profit >= 0 else "📉"
            reply = f"{emoji} Netto-Ergebnis: €{net_profit:.2f}\n• Einnahmen: €{total_income:.2f}\n• Ausgaben: €{total_expenses:.2f}"

        # Lieferant / Vendor (removed "firma"/"şirket" — those go to Firmen now)
        elif any(w in msg for w in ["lieferant", "vendor", "händler", "anbieter", "tedarikçi", "top lieferant"]):
            reply = f"🏢 Top Lieferanten:\n{top_vendors}"

        # ── TIER 4: Page/feature navigation ──

        # Kassenbuch
        elif any(w in msg for w in ["kassenbuch", "bookkeeping", "kasse"]):
            reply = f"📒 Kassenbuch (Bookkeeping):\n• Automatische Synchronisation: Hochgeladene Rechnungen erscheinen automatisch\n• Manuelle Einträge: '+ Eintrag' → Typ, Beschreibung, Lieferant, Betrag, MwSt-Satz, Kategorie, Zahlungsmethode\n• 'Rechnungen sync' Button: Überträgt neue Rechnungen ins Kassenbuch (Duplikate werden übersprungen)\n• Abstimmung (Reconcile): ⬜ klicken → ✅ markiert Eintrag als abgestimmt\n• CSV-Export: Kassenbuch → 'CSV Export'\n• Foto Import: Kassenbuch → 'Foto Import' für handgeschriebene Kassenbücher\n• Bearbeiten & Löschen: Jeder Eintrag einzeln anpassbar\n• 🖨️ PDF Button neben Einträgen mit verknüpfter Rechnung\n\nAktuell: {inv_count} Rechnungen ({len(inv_inc)} Einnahmen, {len(inv_exp)} Ausgaben)"

        # Belege (before generic "rechnung")
        elif any(w in msg for w in ["belege", "beleg", "original", "dokument"]):
            reply = "📎 Belege:\n• Sidebar → 'Belege' Seite\n• Alle hochgeladenen Belege auf einen Blick\n• 'Original ansehen' Button: Vollbild-Vorschau von Bild/PDF\n• 'Download' Button: Original-Datei herunterladen\n• 🖨️ PDF Button: Rechnung als PDF generieren\n• Belege werden in der Datenbank gespeichert (nicht nur als Datei)\n• Suche: Nach Lieferant oder Betrag filtern\n• Tipp: Exportiere regelmäßig als Backup"

        # Rechnung / Invoice (generic)
        elif any(w in msg for w in ["rechnung", "invoice", "faktur", "fatura", "bon", "quittung"]):
            reply = f"🧾 Rechnungen (Invoices):\n• Alle hochgeladenen Belege auf einen Blick: {inv_count} gesamt (€{inv_sum:.2f})\n• Einnahmen: {len(inv_inc)} | Ausgaben: {len(inv_exp)}\n• Suchfunktion: Nach Lieferant, Betrag oder Kategorie suchen\n• Filter: Vendor, Kategorie, Datum (Von/Bis), Status\n• Inline-Bearbeitung: 'Bearbeiten' → Vendor, Betrag, Kategorie, Datum, MwSt ändern\n• Mehrfach löschen: Häkchen setzen → 'X ausgewählte löschen'\n• 🖨️ PDF Button neben jeder Rechnung\n• '+ Rechnung erstellen' für neue Einnahmen\n• Paginierung für große Belegmengen\n\nTipp: 'E-Rechnung' für XML-Import, 'PDF' zum Drucken."

        # Upload
        elif any(w in msg for w in ["upload", "hochladen", "scan", "ocr", "yükle"]):
            reply = "📤 Upload & OCR:\n• Unterstützte Formate: PDF, PNG, JPEG, TIFF, WEBP (max. 5 MB)\n• Einzel- oder Batch-Upload (bis zu 20 Dateien gleichzeitig)\n• OCR erkennt automatisch: Lieferant, Betrag, MwSt-Satz, MwSt-Betrag, Datum\n• Erkannte Kategorien: Lebensmittel, Kraftstoff, Restaurant, Shopping, Transport u.v.m.\n• Über 350 Lieferanten werden automatisch erkannt (Lidl, Amazon, Shell, etc.)\n• Handschrift-Modus für handgeschriebene Belege aktivierbar\n• Einnahme/Ausgabe vor Upload wählbar (Standard: Ausgabe)\n• E-Rechnung (XML) Upload: 'E-Rechnung hochladen' Button\n• QR-Codes auf Rechnungen werden gelesen (EPC/SEPA, Swiss QR)\n• Nach Upload erscheint der Beleg automatisch in Rechnungen UND Kassenbuch"

        # Export
        elif any(w in msg for w in ["export", "excel", "exportieren"]):
            reply = "💾 Export-Optionen:\n• CSV: Komma-getrennte Datei, kompatibel mit Excel, Google Sheets, LibreOffice\n• DATEV: Standard-Format für deutsche Steuerberater — direkt importierbar\n• Excel: .xlsx-Format mit formatierten Spalten\n• JSON: Strukturiertes Datenformat für Entwickler und API-Integration\n• Kassenbuch CSV: Kassenbuch → 'CSV Export' Button\n\nSo geht's:\n1. Gehe zu 'Export'\n2. Wähle optional ein Steuerjahr\n3. Klicke auf CSV, DATEV, Excel oder JSON\n4. Datei wird sofort im Browser heruntergeladen\n\nTipp: Exportierte CSV kann direkt wieder mit 'CSV Import' importiert werden!"

        # CSV
        elif any(w in msg for w in ["csv"]):
            reply = "📄 CSV Funktionen:\n\n• CSV Export (Rechnungen): Export-Seite → 'CSV'\n• CSV Export (Kassenbuch): Kassenbuch → 'CSV Export'\n• CSV Import: Kassenbuch → 'CSV Import'\n\nCSV Format:\nDatum, Lieferant, Rechnungs-Nr., Typ, Betrag, MwSt, MwSt-Satz, Kategorie, Zahlungsart\n• Trennzeichen: Komma oder Semikolon (automatisch erkannt)\n• Spalten: Deutsch oder Englisch\n• Datumsformat: DD.MM.YYYY oder YYYY-MM-DD\n\nTipp: Exportiere erst eine CSV als Vorlage — dann im gleichen Format importieren."

        # Dashboard
        elif any(w in msg for w in ["dashboard", "übersicht", "überblick", "grafik", "chart", "diagramm"]):
            reply = f"📊 Dashboard — Finanzübersicht:\n• Einnahmen: €{total_income:.2f} | Ausgaben: €{total_expenses:.2f}\n• Gewinn: €{net_profit:.2f}\n• MwSt-Saldo: €{vat_balance:.2f}\n• Rechnungen: {inv_count}\n\nFeatures:\n• Steuerschätzung nach deutschem Recht (Progressionsstufen: 0%, 14%, 30%, 42%, 45%)\n• MwSt-Übersicht: Vorsteuer, USt, Saldo\n• Monatliche Auswertung: Einnahmen vs. Ausgaben als Diagramm\n• Kategorien-Verteilung: Wo gibst du am meisten aus?\n• CSV-Export-Button: Alle Rechnungen als CSV\n• 'Alles zurücksetzen': Löscht ALLE Daten (Doppelbestätigung!)"

        # Firmen (generic)
        elif any(w in msg for w in ["firma", "firmen", "unternehmen", "company", "şirket"]):
            reply = "🏢 Firmen verwalten:\n• Sidebar → 'Firmen' Seite\n• Max. 2 Firmen registrieren (Free Plan)\n• Firmenname wird für Einnahme-Erkennung verwendet\n• Upload: Wenn Vendor = deine Firma → automatisch als Einnahme erkannt\n• Firmenname erscheint auf generierten PDFs\n• Firma kann nicht geändert werden (Kontakt: info@autotaxhub.de)"

        # Preise (generic fallback)
        elif any(w in msg for w in ["preis", "fiyat", "was kostet"]):
            reply = "💰 Preise & Pläne:\n• Free: €0/Monat — 50 Rechnungen, 2 Firmen, CSV Export\n• Early Adopter: €10/Monat — 500 Rechnungen, 5 Firmen, DATEV, Excel, PDF\n• Pro: €20/Monat — Unbegrenzt, API, Priority Support\n• Frühe Nutzer behalten ihren Preis dauerhaft!\n• Stripe-Zahlung kommt bald\n\nUpgrade: Sidebar → 'Preise' Seite"

        # PWA / Mobil
        elif any(w in msg for w in ["app", "mobil", "handy", "telefon", "pwa", "installieren", "uygulama"]):
            reply = "📱 Mobile App (PWA):\n• Öffne AutoTax-HUB im Browser auf deinem Handy\n• iPhone: Safari → Teilen → 'Zum Home-Bildschirm'\n• Android: Chrome → Menü → 'App installieren'\n• Funktioniert wie eine native App — kein App Store nötig\n• Belege direkt mit der Kamera hochladen\n• Sidebar: Hamburger-Menü öffnet/schließt Navigation\n• Tipp: Lesezeichen auf dem Home-Bildschirm für schnellen Zugriff"

        # ── TIER 5: Action helpers ──

        # Löschen / Delete
        elif any(w in msg for w in ["lösch", "delete", "entfern", "zurücksetz", "sil", "kaldır", "temizle"]):
            reply = "🗑️ Löschen:\n• Einzeln: Papierkorb-Symbol (✕) neben dem Eintrag\n• Mehrere: Häkchen setzen → 'X ausgewählte löschen' Button\n• Alles zurücksetzen: Dashboard → 'Zurücksetzen'\n  (ACHTUNG: Doppelbestätigung, unwiderruflich! Es gibt kein Undo!)\n\nLöschen funktioniert in Rechnungen UND Kassenbuch.\nGelöschte Einträge sind sofort weg — vorher exportieren empfohlen."

        # Passwort / Login
        elif any(w in msg for w in ["passwort", "password", "şifre", "kennwort", "login", "anmeld", "registrier", "konto"]):
            reply = "🔐 Konto & Sicherheit:\n• Passwort: Min. 8 Zeichen, 1 Großbuchstabe, 1 Zahl (Beispiel: MeinPasswort1)\n• Login: E-Mail + Passwort\n• Token: Automatische Erneuerung (1h Access, 7 Tage Refresh)\n• Registrierung: Auf der Login-Seite 'Registrieren' klicken\n• Passwort ändern: Profil-Menü oben rechts → 'Passwort ändern'"

        # Sync / Synchronisieren
        elif any(w in msg for w in ["sync", "synchron", "senkron"]):
            reply = "🔄 Synchronisation:\n• Upload → Beleg erscheint automatisch in Rechnungen + Kassenbuch\n• Kassenbuch → 'Rechnungen sync' synchronisiert fehlende Einträge\n• Rechnungen → 'Kassenbuch sync' synchronisiert in beide Richtungen\n• Duplikate werden automatisch erkannt und übersprungen\n• Reverse sync: Manuelle Kassenbuch-Einträge → Rechnungen"

        # Reconcile / Abstimmen
        elif any(w in msg for w in ["reconcil", "abstimm", "häkchen", "checkbox"]):
            reply = "✅ Abstimmung (Reconcile):\n• Kassenbuch → Klicke ⬜ neben einem Eintrag → wird ✅\n• Markiert den Eintrag als 'abgestimmt mit Kontoauszug'\n• Hilft beim Abgleich mit Bankkontoauszügen\n• Kann jederzeit rückgängig gemacht werden (✅ → ⬜)"

        # QR Code
        elif any(w in msg for w in ["qr", "barcode"]):
            reply = "📱 QR-Code Erkennung:\n• QR-Codes auf Rechnungen werden automatisch gelesen\n• Unterstützt: EPC/SEPA (GiroCode), Swiss QR, ZUGFeRD\n• Extrahiert: Firma, IBAN, Betrag, Referenz\n• QR-Daten überschreiben OCR wenn verfügbar (genauer)\n• Tipp: QR-Code muss gut lesbar im Bild sein"

        # Foto / Bild Qualität
        elif any(w in msg for w in ["foto", "qualität", "unscharf", "dunkel", "yamuk", "blurry", "bild"]):
            reply = "📸 Foto-Tipps für bessere Erkennung:\n• Gute Beleuchtung — kein Schatten auf dem Beleg\n• Gerade fotografieren — nicht schief\n• Gesamten Beleg im Bild — nichts abgeschnitten\n• Original-Foto verwenden (nicht WhatsApp-komprimiert)\n• PDF ist besser als Foto (wenn verfügbar)\n• Handschrift-Modus für handgeschriebene Belege aktivieren\n• Tipp: Kamera-App statt Screenshot verwenden"

        # Einnahme / Ausgabe
        elif any(w in msg for w in ["einnahme oder ausgabe", "ausgabe oder einnahme", "einnahme ausgabe"]):
            reply = "📈 Einnahme / Ausgabe:\n• Upload-Seite: Vor dem Hochladen 'Ausgabe (Gider)' oder 'Einnahme (Gelir)' wählen\n• Standard ist Ausgabe — für Einnahmen den grünen Button klicken\n• Kassenbuch: '+ Eintrag' → Typ 'Einnahme' oder 'Ausgabe' im Formular wählen\n• Dashboard zeigt Einnahmen (grün) und Ausgaben (rot) getrennt\n• Gewinn = Einnahmen minus Ausgaben\n• Auto-Erkennung: Wenn Vendor = deine Firma → automatisch Einnahme"

        # Eintragen / Hinzufügen
        elif any(w in msg for w in ["eintragen", "hinzufügen", "eingeben", "ekle", "kaydet", "erfassen"]):
            reply = "✏️ Eintrag erstellen:\n• Upload → Beleg hochladen (OCR erkennt automatisch alles)\n• Kassenbuch → '+ Eintrag' Button → Formular ausfüllen:\n  Typ, Beschreibung, Lieferant, Betrag, MwSt-Satz, Kategorie, Zahlungsmethode\n• Rechnungen → '+ Rechnung erstellen' für Einnahmen\n• CSV Import: Kassenbuch → 'CSV Import' für Masseneingabe\n• Foto Import: Kassenbuch → 'Foto Import' für handgeschriebene Bücher\n• Beide Wege erstellen Einträge in Rechnungen UND Kassenbuch"

        # Suche / Finden
        elif any(w in msg for w in ["such", "find", "wo ist", "wo sind", "finden", "ara", "bul", "nerede", "search", "where"]):
            reply = "🔍 Suche:\n• Rechnungen → Suchfeld oben (sucht in Vendor, OCR-Text, Kategorie)\n• Mehrere Wörter möglich: z.B. 'Lidl Dezember'\n• Filter: Vendor, Kategorie, Datum (Von/Bis), Status\n• Kassenbuch → Eigenes Suchfeld\n• Belege → Suchfeld für Lieferant/Betrag\n• AI Chat: Frag mich z.B. einen Lieferanten-Namen wie 'Lidl'"

        # Bearbeiten / Ändern
        elif any(w in msg for w in ["bearbeit", "änder", "korrigier", "edit", "düzenle", "değiştir"]):
            reply = "✏️ Bearbeiten:\n• Rechnungen → 'Bearbeiten' neben dem Eintrag\n• Kassenbuch → 'Bearbeiten' neben dem Eintrag\n• Änderbare Felder: Vendor, Betrag, MwSt-Betrag, MwSt-Satz, Kategorie, Datum, Rechnungs-Nr., Zahlungsart\n• Tipp: Wenn OCR falsch erkannt hat → hier korrigieren"

        # Datum / Date
        elif any(w in msg for w in ["datum", "tarih", "zeitraum", "monat", "jahr"]):
            reply = "📅 Datum-Filter:\n• Rechnungen → Von/Bis Felder nutzen (nur Kalender-Auswahl)\n• Unterstützte Formate: DD.MM.YYYY, YYYY-MM-DD\n• Monatsansicht: Dashboard zeigt monatliche Auswertung als Diagramm\n• Export: Nach Steuerjahr filterbar\n• Wenn OCR Datum nicht erkennt → heutiges Datum als Fallback"

        # Download / Herunterladen (generic fallback)
        elif any(w in msg for w in ["download", "herunterladen"]):
            reply = "📥 Download-Optionen:\n• Rechnung als PDF: 🖨️ PDF Button neben jeder Rechnung\n• Original-Beleg: Belege → 'Download' Button\n• Daten exportieren: Export-Seite → CSV, DATEV, Excel, JSON\n• Kassenbuch CSV: Kassenbuch → 'CSV Export'"

        # ── TIER 6: Vendor search fallback ──
        # Vendor search — if no keyword matched, try searching vendor names
        if reply is None:
            vendor_results = [i for i in invoices if msg in (i.vendor or "").lower()]
            if not vendor_results:
                vendor_results = db.query(Invoice).filter(Invoice.user_id == user["sub"], Invoice.vendor.ilike(f"%{msg}%")).all()
            if vendor_results:
                vr_total = sum(safe_float(i.total_amount) for i in vendor_results)
                vr_vat = sum(safe_float(i.vat_amount) for i in vendor_results)
                latest = safe_date_str(vendor_results[0].date) if vendor_results[0].date else "unbekannt"
                reply = f"🔍 Für '{msg.title()}' habe ich {len(vendor_results)} Rechnung(en) gefunden:\n• Gesamtbetrag: €{vr_total:.2f}\n• MwSt: €{vr_vat:.2f}\n• Letzte Rechnung: {latest}"
            else:
                reply = f"Das habe ich nicht ganz verstanden. Versuche z.B.:\n• 'Wie viele Rechnungen?'\n• 'MwSt Übersicht'\n• 'Gewinn'\n• Einen Lieferanten-Namen (z.B. 'Lidl')\n• 'Hilfe' für alle Themen\n\nAktuell: {inv_count} Rechnungen, €{net_profit:.2f} Gewinn"

        return {"reply": reply}
    except Exception:
        logger.exception("Chat failed")
        return {"reply": "Entschuldigung, ein Fehler ist aufgetreten. Bitte versuche es erneut."}
    finally:
        db.close()


# ============================================================
# EXPORT: DATEV / CSV / EXCEL / JSON
# ============================================================

@app.get("/export/csv")
def export_csv(year: int = Query(None), user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"]).all()
        buf = io.StringIO()
        buf.write("# HINWEIS: Automatisch erstellt von AutoTax-HUB. Alle Daten vor Verwendung pruefen. Keine Steuerberatung.\n")
        buf.write("Datum,Lieferant,Rechnungs-Nr.,Typ,Betrag,MwSt,MwSt-Satz,Kategorie,Zahlungsart\n")
        for i in invoices:
            d = safe_date_str(i.date)
            if year and not d.startswith(str(year)):
                continue
            vendor = (i.vendor or "").replace('"', '""')
            buf.write(f'{d},"{vendor}",{safe_str(i.invoice_number)},{safe_invoice_type(i.invoice_type)},{safe_float(i.total_amount):.2f},{safe_float(i.vat_amount):.2f},{safe_vat_rate(i.vat_rate)},{safe_category(i.category)},{safe_str(i.payment_method)}\n')
        buf.seek(0)
        return StreamingResponse(buf, media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=autotax_csv_{year or 'all'}.csv"})
    except Exception:
        logger.exception("CSV export failed")
        err(500, "Export failed")
    finally:
        db.close()


@app.get("/export/datev")
def export_datev(year: int = Query(None), user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"]).all()
        buf = io.StringIO()
        buf.write("# HINWEIS: AutoTax-HUB - Alle Daten pruefen. Keine Steuerberatung.\n")
        buf.write("Umsatz;Soll/Haben;Konto;Gegenkonto;BU;Belegdatum;Buchungstext;USt\n")
        for i in invoices:
            d = safe_date_str(i.date)
            if year and not d.startswith(str(year)):
                continue
            sh = "S" if safe_invoice_type(i.invoice_type) == "expense" else "H"
            date_str = ""
            parts = d.split("-")
            if len(parts) == 3:
                date_str = f"{parts[2]}{parts[1]}"
            amt = f"{safe_float(i.total_amount):.2f}".replace(".", ",")
            vendor = (i.vendor or "").replace(";", " ")
            vat = (i.vat_rate or "0%").replace("%", "")
            cat = safe_category(i.category)
            konto = _DATEV_KONTO_MAP.get(cat, "6800") if sh == "S" else _DATEV_KONTO_MAP_INCOME.get(cat, "8400")
            buf.write(f"{amt};{sh};{konto};1200;{vat};{date_str};{vendor};{vat}\n")
        buf.seek(0)
        return StreamingResponse(buf, media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=autotax_datev_{year or 'all'}.csv"})
    except Exception:
        logger.exception("DATEV export failed")
        err(500, "Export failed")
    finally:
        db.close()


@app.get("/export/excel")
def export_excel(year: int = Query(None), user: dict = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    db = SessionLocal()
    try:
        invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"]).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "AutoTax Export"
        # Disclaimer row
        disc = ws.cell(row=1, column=1, value="HINWEIS: Automatisch erstellt von AutoTax-HUB. Alle Daten vor Verwendung prüfen. Keine Steuerberatung.")
        disc.font = Font(italic=True, color="F59E0B")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        headers = ["Datum", "Lieferant", "Rechnungs-Nr.", "Typ", "Betrag", "MwSt", "MwSt-Satz", "Kategorie", "Zahlungsart"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row = 3
        for i in invoices:
            d = safe_date_str(i.date)
            if year and not d.startswith(str(year)):
                continue
            ws.cell(row=row, column=1, value=d)
            ws.cell(row=row, column=2, value=safe_vendor(i.vendor))
            ws.cell(row=row, column=3, value=safe_str(i.invoice_number))
            ws.cell(row=row, column=4, value=safe_invoice_type(i.invoice_type))
            c5 = ws.cell(row=row, column=5, value=safe_float(i.total_amount))
            c5.number_format = '#,##0.00'
            c6 = ws.cell(row=row, column=6, value=safe_float(i.vat_amount))
            c6.number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=safe_vat_rate(i.vat_rate))
            ws.cell(row=row, column=8, value=safe_category(i.category))
            ws.cell(row=row, column=9, value=safe_str(i.payment_method))
            row += 1
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 18
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=autotax_{year or 'all'}.xlsx"})
    except Exception:
        logger.exception("Excel export failed")
        err(500, "Export failed")
    finally:
        db.close()


@app.get("/export/json")
def export_json(year: int = Query(None), user: dict = Depends(get_current_user)):
    import json as json_lib
    db = SessionLocal()
    try:
        invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"]).all()
        data = []
        for i in invoices:
            d = safe_date_str(i.date)
            if year and not d.startswith(str(year)):
                continue
            data.append(invoice_to_dict(i))
        buf = io.StringIO()
        json_lib.dump(data, buf, indent=2, ensure_ascii=False)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/json", headers={"Content-Disposition": f"attachment; filename=autotax_{year or 'all'}.json"})
    except Exception:
        logger.exception("JSON export failed")
        err(500, "Export failed")
    finally:
        db.close()



# --- ADDED START: Soft delete restore + trash endpoints ---

@app.get("/invoices/deleted")
def list_deleted_invoices(user: dict = Depends(get_current_user)):
    """List all soft-deleted invoices (trash)."""
    db = SessionLocal()
    try:
        invs = db.query(Invoice).filter(Invoice.user_id == user["sub"], Invoice.is_deleted == True).order_by(Invoice.deleted_at.desc()).all()
        return {"success": True, "items": [invoice_to_dict(i) | {"deleted_at": i.deleted_at.strftime("%Y-%m-%dT%H:%M:%S") if i.deleted_at else ""} for i in invs], "total": len(invs)}
    finally:
        db.close()


@app.post("/invoices/{invoice_id}/restore")
def restore_invoice(invoice_id: int, user: dict = Depends(get_current_user)):
    """Restore a soft-deleted invoice."""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        inv.is_deleted = False
        inv.deleted_at = None
        logger.info("Restored invoice %d", invoice_id)
        # Also restore linked cash entry
        linked = db.query(CashEntry).filter(CashEntry.invoice_id == invoice_id, CashEntry.user_id == user["sub"]).first()
        if linked:
            linked.is_deleted = False
            linked.deleted_at = None
            logger.info("Restored linked cash entry for invoice %d", invoice_id)
        db.commit()
        return {"success": True, "restored": invoice_id}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Restore invoice failed")
        err(500, "Restore failed")
    finally:
        db.close()


@app.delete("/invoices/{invoice_id}/permanent")
def permanent_delete_invoice(invoice_id: int, user: dict = Depends(get_current_user)):
    """Permanently delete an invoice (from trash)."""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        # Also delete linked cash entry
        db.query(CashEntry).filter(CashEntry.invoice_id == invoice_id, CashEntry.user_id == user["sub"]).delete()
        db.delete(inv)
        db.commit()
        logger.info("Permanent delete: invoice %d", invoice_id)
        return {"success": True, "deleted": invoice_id}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Permanent delete failed")
        err(500, "Delete failed")
    finally:
        db.close()

# --- ADDED END ---

# Build trigger: 2026-03-24
# force deploy 2
